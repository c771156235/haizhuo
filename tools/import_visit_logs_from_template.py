# -*- coding: utf-8 -*-
"""
从「拜访日志模板 Excel」批量导入拜访日志。

设计目标：
1) 模板结构与 tools/generate_visit_log_template_from_work_orders.py 的导出一致。
2) 每行通过「工单编号」定位工单，创建一条拜访日志（默认一单一日志）。
3) 尽量复用线上创建拜访日志时的快照与状态流转逻辑。

示例：
  cd backend
  python ..\\tools\\import_visit_logs_from_template.py -i f:\\FDE_project\\拜访日志模板_已填写.xlsx --dry-run

  python ..\\tools\\import_visit_logs_from_template.py -i f:\\FDE_project\\拜访日志模板_已填写.xlsx

  python ..\\tools\\import_visit_logs_from_template.py -i f:\\FDE_project\\拜访日志模板_已填写.xlsx --allow-update-existing
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

EXPECTED_HEADERS = [
    "工单编号",
    "任务名称",
    "客户单位",
    "客户拜访地址",
    "客户经理",
    "客户经理联系方式",
    "组别",
    "所属销售单位",
    "行业",
    "企业类型",
    "陪跑人员",
    "创建人",
    "拜访日期",
    "拜访对象职位",
    "拜访对象权限",
    "是否有线索",
    "线索对应产品",
    "当前阶段",
    "阶段人员投入与时长",
    "推进进展",
    "推进要求",
    "是否定开",
    "定开要求",
    "预估金额（万元）",
    "客户是否梳理过需求场景",
    "需求场景分类",
    "拜访内容",
    "备注",
    "创建时间",
]


def _setup_backend_path() -> None:
    root = Path(__file__).resolve().parent.parent
    backend = root / "backend"
    if not backend.is_dir():
        raise SystemExit(f"未找到 backend 目录: {backend}")
    sys.path.insert(0, str(backend))
    os.chdir(backend)


def s(value: Any) -> Optional[str]:
    if value is None:
        return None
    t = str(value).strip()
    return t if t else None


def parse_bool(value: Any, field_name: str) -> bool:
    if value is None or str(value).strip() == "":
        return False
    t = str(value).strip().lower()
    if t in ("是", "y", "yes", "true", "1"):
        return True
    if t in ("否", "n", "no", "false", "0"):
        return False
    raise ValueError(f"{field_name} 仅支持：是/否（或 true/false, 1/0）")


def parse_date(value: Any, field_name: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    t = s(value)
    if not t:
        raise ValueError(f"{field_name} 不能为空")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(t, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name} 格式错误，示例：2026-01-28")


def parse_optional_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    t = s(value)
    if not t:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            parsed = datetime.strptime(t, fmt)
            if fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                return datetime.combine(parsed.date(), datetime.min.time())
            return parsed
        except ValueError:
            continue
    raise ValueError(f"创建时间格式错误：{t}")


def parse_requirement_scenario_category(value: Any) -> Optional[str]:
    """
    导入到 requirement_scenario_category（DB 存 JSON 数组字符串）。
    支持:
    - 空
    - JSON数组字符串
    - 逗号/顿号分隔文本
    """
    from app.utils.visit_log_requirement_scenario_utils import (
        VISIT_LOG_REQUIREMENT_SCENARIO_CATEGORY_LABELS,
    )

    raw = s(value)
    if not raw:
        return None

    # 1) 若本身是 JSON 数组
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            labels = [str(x).strip() for x in parsed if str(x).strip()]
            if not labels:
                return None
            invalid = [x for x in labels if x not in VISIT_LOG_REQUIREMENT_SCENARIO_CATEGORY_LABELS]
            if invalid:
                raise ValueError(f"需求场景分类存在无效值: {', '.join(invalid)}")
            return json.dumps(labels, ensure_ascii=False)
    except json.JSONDecodeError:
        pass

    # 2) 按导出文本 "A, B" / "A，B" / "A、B"
    parts = [p.strip() for p in re.split(r"[,，、]", raw) if p and p.strip()]
    if not parts:
        return None
    invalid = [x for x in parts if x not in VISIT_LOG_REQUIREMENT_SCENARIO_CATEGORY_LABELS]
    if invalid:
        raise ValueError(f"需求场景分类存在无效值: {', '.join(invalid)}")
    return json.dumps(parts, ensure_ascii=False)


def parse_stage_effort_breakdown(value: Any, current_stage: Optional[str]) -> Optional[str]:
    """
    导入到 stage_effort_breakdown（DB 存 JSON 字符串）。
    支持两种输入：
    - JSON数组
    - 导出可读文本：需求排摸：人员2人，时长3天；标品试用：人员1人，时长5天
    """
    from app.schemas.visit_log import normalize_stage_effort_breakdown_json

    raw = s(value)
    if not raw:
        return None

    if raw.startswith("["):
        return normalize_stage_effort_breakdown_json(current_stage, raw)

    # 尝试解析导出展示文本
    items = [x.strip() for x in re.split(r"[；;]", raw) if x.strip()]
    parsed: list[dict[str, Any]] = []
    for item in items:
        # 允许：<子阶段>：人员X人，时长Y天（X/Y 可为空或小数）
        m = re.match(
            r"^\s*(?P<sub>[^：:]+)\s*[：:]\s*人员\s*(?P<people>-?\d+(?:\.\d+)?)?\s*人?\s*[,，]\s*时长\s*(?P<days>-?\d+(?:\.\d+)?)?\s*天?\s*$",
            item,
        )
        if not m:
            raise ValueError(f"阶段人员投入与时长格式无法解析：{item}")
        pe = m.group("people")
        da = m.group("days")
        parsed.append(
            {
                "sub_phase": m.group("sub").strip(),
                "people": float(pe) if pe is not None else None,
                "days": float(da) if da is not None else None,
            }
        )
    return normalize_stage_effort_breakdown_json(
        current_stage,
        json.dumps(parsed, ensure_ascii=False),
    )


def row_is_empty(row_values: dict[str, Any]) -> bool:
    for v in row_values.values():
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def main() -> None:
    _setup_backend_path()

    from sqlalchemy.orm import joinedload

    from app.core.workflow import check_and_update_task_completion_status
    from app.models.task import Task
    from app.models.visit_log import VisitLog
    from app.models.work_order import WorkOrder, WorkOrderStatus
    from app.database import SessionLocal
    from app.schemas.visit_log import (
        VISIT_LOG_CURRENT_STAGE_OPTIONS,
        VISIT_LOG_DECISION_AUTHORITY_OPTIONS,
        VISIT_LOG_ENTERPRISE_TYPES,
    )
    from app.api.visit_logs import (
        resolve_visit_log_customer_unit_snapshot,
        resolve_visit_log_detail_contact_snapshot,
        resolve_visit_log_sales_unit_snapshot,
    )
    from app.api.work_orders import get_fde_group_name_for_team_leader
    from app.services.work_order_completion import set_work_order_completed_at_if_missing

    parser = argparse.ArgumentParser(description="导入拜访日志模板 Excel 到数据库")
    parser.add_argument("-i", "--input", required=True, help="输入模板 .xlsx 路径")
    parser.add_argument(
        "--allow-update-existing",
        action="store_true",
        help="若工单已有拜访日志，则改为更新该条日志（默认：报错并跳过）",
    )
    parser.add_argument(
        "--allow-any-status",
        action="store_true",
        help="默认仅允许 accepted/in_progress 工单导入；开启后放宽状态限制",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="仅校验和预演，不写入数据库",
    )
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if not input_path.exists():
        raise SystemExit(f"文件不存在: {input_path}")

    wb = load_workbook(str(input_path), data_only=True)
    ws = wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    actual_headers = [str(x).strip() if x is not None else "" for x in header_cells]
    expected_prefix = EXPECTED_HEADERS
    if actual_headers[: len(expected_prefix)] != expected_prefix:
        raise SystemExit(
            "表头不匹配，请使用系统模板。\n"
            f"期望前{len(expected_prefix)}列: {expected_prefix}\n"
            f"实际前{len(expected_prefix)}列: {actual_headers[:len(expected_prefix)]}"
        )
    header_to_idx = {h: i for i, h in enumerate(actual_headers)}

    db = SessionLocal()
    success = 0
    skipped = 0
    failed = 0
    errors: list[str] = []
    touched_task_ids: set[int] = set()
    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_values = {h: row[idx] if idx < len(row) else None for h, idx in header_to_idx.items()}
            if row_is_empty(row_values):
                continue

            try:
                work_order_no = s(row_values["工单编号"])
                if not work_order_no:
                    raise ValueError("工单编号不能为空")

                work_order = (
                    db.query(WorkOrder)
                    .options(
                        joinedload(WorkOrder.task),
                        joinedload(WorkOrder.detail_requirement),
                        joinedload(WorkOrder.member),
                        joinedload(WorkOrder.visit_logs),
                    )
                    .filter(WorkOrder.work_order_no == work_order_no)
                    .first()
                )
                if not work_order:
                    raise ValueError(f"工单不存在：{work_order_no}")

                if (
                    not args.allow_any_status
                    and work_order.status not in (WorkOrderStatus.ACCEPTED.value, WorkOrderStatus.IN_PROGRESS.value)
                ):
                    raise ValueError(
                        f"工单状态不允许导入：{work_order.status}（默认仅 accepted/in_progress，可用 --allow-any-status 放宽）"
                    )

                if not work_order.member_id:
                    raise ValueError("工单未分配成员，无法确定拜访日志创建人")

                existing = db.query(VisitLog).filter(VisitLog.work_order_id == work_order.id).first()
                if existing and not args.allow_update_existing:
                    skipped += 1
                    errors.append(f"第{row_idx}行：工单 {work_order_no} 已有拜访日志(id={existing.id})，已跳过")
                    continue

                visit_date = parse_date(row_values["拜访日期"], "拜访日期")
                visit_content = s(row_values["拜访内容"])
                if not visit_content:
                    raise ValueError("拜访内容不能为空")

                industry = s(row_values["行业"])
                if not industry:
                    raise ValueError("行业不能为空")

                enterprise_type = s(row_values["企业类型"])
                if not enterprise_type:
                    raise ValueError("企业类型不能为空")
                if enterprise_type not in VISIT_LOG_ENTERPRISE_TYPES:
                    raise ValueError(f"企业类型无效：{enterprise_type}，允许值：{', '.join(VISIT_LOG_ENTERPRISE_TYPES)}")

                has_decision_authority = s(row_values["拜访对象权限"])
                if has_decision_authority and has_decision_authority not in VISIT_LOG_DECISION_AUTHORITY_OPTIONS:
                    raise ValueError(
                        f"拜访对象权限无效：{has_decision_authority}，允许值：{', '.join(VISIT_LOG_DECISION_AUTHORITY_OPTIONS)}"
                    )

                has_clue = parse_bool(row_values["是否有线索"], "是否有线索")
                is_customized_development = parse_bool(row_values["是否定开"], "是否定开")
                has_requirement_scenario_sorted = parse_bool(
                    row_values["客户是否梳理过需求场景"], "客户是否梳理过需求场景"
                )

                current_stage = s(row_values["当前阶段"])
                if current_stage and current_stage not in VISIT_LOG_CURRENT_STAGE_OPTIONS:
                    raise ValueError(
                        f"当前阶段无效：{current_stage}，允许值：{', '.join(VISIT_LOG_CURRENT_STAGE_OPTIONS)}"
                    )

                stage_effort_breakdown = parse_stage_effort_breakdown(
                    row_values["阶段人员投入与时长"], current_stage
                )
                requirement_scenario_category = parse_requirement_scenario_category(
                    row_values["需求场景分类"]
                )

                customized_development_requirements = s(row_values["定开要求"])
                if is_customized_development and not customized_development_requirements:
                    raise ValueError("是否定开=是 时，定开要求不能为空")

                created_at = parse_optional_datetime(row_values["创建时间"])

                fde_group_name = get_fde_group_name_for_team_leader(db, work_order.team_leader_id)
                customer_unit_snapshot = resolve_visit_log_customer_unit_snapshot(work_order)
                sales_unit_snapshot = resolve_visit_log_sales_unit_snapshot(work_order)
                addr_snap, mgr_name_snap, mgr_contact_snap = resolve_visit_log_detail_contact_snapshot(work_order)

                payload = dict(
                    work_order_id=work_order.id,
                    member_id=work_order.member_id,
                    visit_date=visit_date,
                    visit_content=visit_content,
                    remark=s(row_values["备注"]),
                    visit_object_position=s(row_values["拜访对象职位"]),
                    has_decision_authority=has_decision_authority,
                    has_clue=has_clue,
                    clue_related_products=s(row_values["线索对应产品"]),
                    current_stage=current_stage,
                    stage_effort_breakdown=stage_effort_breakdown,
                    promotion_progress=s(row_values["推进进展"]),
                    promotion_requirements=s(row_values["推进要求"]),
                    is_customized_development=is_customized_development,
                    customized_development_requirements=customized_development_requirements,
                    project_amount=s(row_values["预估金额（万元）"]),
                    has_requirement_scenario_sorted=has_requirement_scenario_sorted,
                    requirement_scenario_category=requirement_scenario_category,
                    industry=industry,
                    enterprise_type=enterprise_type,
                    escort_staff=s(row_values["陪跑人员"]),
                    sales_unit=sales_unit_snapshot or None,
                    group_name=fde_group_name,
                    customer_unit=customer_unit_snapshot,
                    customer_visit_address=addr_snap,
                    customer_manager_name=mgr_name_snap,
                    customer_manager_contact=mgr_contact_snap,
                )

                if existing and args.allow_update_existing:
                    for k, v in payload.items():
                        setattr(existing, k, v)
                    target = existing
                else:
                    target = VisitLog(**payload)
                    db.add(target)

                if created_at is not None:
                    target.created_at = created_at

                # 与线上逻辑保持一致：创建日志后工单进入已拜访
                if work_order.status in (WorkOrderStatus.ACCEPTED.value, WorkOrderStatus.IN_PROGRESS.value):
                    work_order.status = WorkOrderStatus.COMPLETED.value
                    set_work_order_completed_at_if_missing(work_order)

                if work_order.task_id:
                    touched_task_ids.add(work_order.task_id)

                if args.dry_run:
                    db.flush()
                    db.rollback()
                else:
                    db.flush()
                    success += 1
            except Exception as exc:
                failed += 1
                db.rollback()
                errors.append(f"第{row_idx}行：{exc}")
                continue

        if args.dry_run:
            print("dry-run 完成：未写入数据库")
        else:
            # 统一更新受影响任务状态
            for task_id in touched_task_ids:
                task_obj = db.query(Task).filter(Task.id == task_id).first()
                if task_obj:
                    check_and_update_task_completion_status(db, task_obj)
            db.commit()

        print(
            f"导入结束：成功 {success} 行，跳过 {skipped} 行，失败 {failed} 行。"
            f" dry_run={args.dry_run}, allow_update_existing={args.allow_update_existing}"
        )
        if errors:
            print("\n明细：")
            for e in errors:
                print(f"- {e}")
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    main()
