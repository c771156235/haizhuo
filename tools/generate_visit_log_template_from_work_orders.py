# -*- coding: utf-8 -*-
"""
从数据库中的工单列表生成「拜访日志」Excel 模板。

表头与列宽与系统导出 `export_visit_logs_to_excel`（线索维护 / 拜访日志导出）一致；
预填字段与创建拜访日志时工单快照逻辑一致（客户单位、销售单位、地址、客户经理、组别、任务名称、行业、创建人等）。

用法（在仓库根目录或 backend 目录均可）:
  cd backend
  python ..\\tools\\generate_visit_log_template_from_work_orders.py -o visit_log_template.xlsx

  python ..\\tools\\generate_visit_log_template_from_work_orders.py -o out.xlsx \\
      --status accepted,in_progress --skip-with-visit-log

  python ..\\tools\\generate_visit_log_template_from_work_orders.py -o out.xlsx --task-id 12
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 与 backend/app/api/export.py 中 export_visit_logs_to_excel 保持一致
VISIT_LOG_EXPORT_HEADERS = [
    "工单编号",
    "任务名称",
    "客户单位",
    "客户拜访地址",
    "客户经理",
    "客户经理联系方式",
    "组别",
    "所属销售单位",
    "行业",
    "企业类型",
    "陪跑人员",
    "创建人",
    "拜访日期",
    "拜访对象职位",
    "拜访对象权限",
    "是否有线索",
    "线索对应产品",
    "当前阶段",
    "阶段人员投入与时长",
    "推进进展",
    "推进要求",
    "是否定开",
    "定开要求",
    "预估金额（万元）",
    "客户是否梳理过需求场景",
    "需求场景分类",
    "拜访内容",
    "备注",
    "创建时间",
]

VISIT_LOG_EXPORT_COLUMN_WIDTHS = [
    20,
    28,
    18,
    24,
    12,
    14,
    14,
    18,
    14,
    12,
    14,
    15,
    12,
    20,
    18,
    12,
    28,
    16,
    36,
    24,
    24,
    12,
    18,
    28,
    32,
    16,
    24,
    20,
    20,
]


def _setup_backend_path() -> None:
    root = Path(__file__).resolve().parent.parent
    backend = root / "backend"
    if not backend.is_dir():
        raise SystemExit(f"未找到 backend 目录: {backend}")
    sys.path.insert(0, str(backend))
    os.chdir(backend)


def resolve_industry_type(work_order) -> str:
    """与工单详情展示一致：详细需求行业优先，否则任务级。"""
    if work_order.detail_requirement and work_order.detail_requirement.industry_type:
        return (work_order.detail_requirement.industry_type or "").strip()
    if work_order.task and work_order.task.industry_type:
        return (work_order.task.industry_type or "").strip()
    return ""


def build_row_for_work_order(db, work_order) -> list:
    """一行数据：与导出列顺序一致；可预填的来自工单链，其余留空。"""
    from app.api.visit_logs import (
        resolve_visit_log_customer_unit_snapshot,
        resolve_visit_log_detail_contact_snapshot,
        resolve_visit_log_sales_unit_snapshot,
    )
    from app.api.work_orders import get_fde_group_name_for_team_leader

    customer_unit = resolve_visit_log_customer_unit_snapshot(work_order) or ""
    addr, mgr_name, mgr_contact = resolve_visit_log_detail_contact_snapshot(work_order)
    addr = addr or ""
    mgr_name = mgr_name or ""
    mgr_contact = mgr_contact or ""
    sales_unit = resolve_visit_log_sales_unit_snapshot(work_order) or ""
    group_name = get_fde_group_name_for_team_leader(db, work_order.team_leader_id) or ""
    task_name = ""
    if work_order.task:
        task_name = work_order.task.task_name or ""
    industry = resolve_industry_type(work_order)
    creator = ""
    if work_order.member:
        creator = work_order.member.real_name or ""

    # 以下列留给用户填写（与线上新建拜访日志一致）
    empty = ""
    return [
        work_order.work_order_no or "",
        task_name,
        customer_unit,
        addr,
        mgr_name,
        mgr_contact,
        group_name,
        sales_unit,
        industry,
        empty,  # 企业类型
        empty,  # 陪跑人员
        creator,
        empty,  # 拜访日期
        empty,  # 拜访对象职位
        empty,  # 拜访对象权限
        empty,  # 是否有线索
        empty,  # 线索对应产品
        empty,  # 当前阶段
        empty,  # 阶段人员投入与时长
        empty,  # 推进进展
        empty,  # 推进要求
        empty,  # 是否定开
        empty,  # 定开要求
        empty,  # 预估金额（万元）
        empty,  # 客户是否梳理过需求场景
        empty,  # 需求场景分类
        empty,  # 拜访内容
        empty,  # 备注
        empty,  # 创建时间（导入时由系统生成亦可留空）
    ]


def parse_statuses(raw: str | None) -> list[str]:
    if not raw or not raw.strip():
        return ["accepted", "in_progress"]
    return [s.strip() for s in raw.split(",") if s.strip()]


def main() -> None:
    _setup_backend_path()

    from sqlalchemy.orm import joinedload

    from app.database import SessionLocal
    from app.models.visit_log import VisitLog
    from app.models.work_order import WorkOrder, WorkOrderStatus

    parser = argparse.ArgumentParser(description="从工单生成拜访日志 Excel 模板")
    parser.add_argument(
        "-o",
        "--output",
        required=True,
        help="输出 .xlsx 路径",
    )
    parser.add_argument(
        "--status",
        default="accepted,in_progress",
        help="工单状态过滤，逗号分隔，默认 accepted,in_progress",
    )
    parser.add_argument(
        "--include-with-visit-log",
        action="store_true",
        help="默认会排除已有拜访日志的工单；指定本项则不过滤（仍导出这些工单行）",
    )
    parser.add_argument("--task-id", type=int, default=None, help="仅指定任务 ID 的工单")
    parser.add_argument("--member-id", type=int, default=None, help="仅指定承接成员用户 ID")
    parser.add_argument(
        "--team-leader-id", type=int, default=None, help="仅指定组长用户 ID"
    )
    args = parser.parse_args()

    skip_vl = not args.include_with_visit_log
    statuses = parse_statuses(args.status)

    db = SessionLocal()
    try:
        q = (
            db.query(WorkOrder)
            .options(
                joinedload(WorkOrder.task),
                joinedload(WorkOrder.detail_requirement),
                joinedload(WorkOrder.member),
            )
            .filter(WorkOrder.status.in_(statuses))
        )
        if args.task_id is not None:
            q = q.filter(WorkOrder.task_id == args.task_id)
        if args.member_id is not None:
            q = q.filter(WorkOrder.member_id == args.member_id)
        if args.team_leader_id is not None:
            q = q.filter(WorkOrder.team_leader_id == args.team_leader_id)

        if skip_vl:
            wo_with_vl = {r[0] for r in db.query(VisitLog.work_order_id).distinct().all()}
            if wo_with_vl:
                q = q.filter(~WorkOrder.id.in_(wo_with_vl))

        work_orders = q.order_by(WorkOrder.id.asc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "拜访日志列表"

        ws.append(VISIT_LOG_EXPORT_HEADERS)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for wo in work_orders:
            ws.append(build_row_for_work_order(db, wo))

        for i, width in enumerate(VISIT_LOG_EXPORT_COLUMN_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        out_path = Path(args.output).resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out_path))
        print(
            f"已生成: {out_path} ，共 {len(work_orders)} 行工单数据"
            f"（状态 in {statuses}，skip_with_visit_log={skip_vl}）"
        )
    finally:
        db.close()


if __name__ == "__main__":
    main()
