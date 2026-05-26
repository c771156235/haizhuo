"""
数据导出 API
"""
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from datetime import datetime
import io
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.fonts import addMapping
import os
import platform
import glob
import logging
import subprocess
import sys
import re

# 配置日志
logger = logging.getLogger(__name__)

from app.database import get_db
from app.models.user import User, UserRole
from app.models.task import Task, TaskDetailRequirement
from app.models.work_order import WorkOrder
from app.models.opportunity import Opportunity
from app.models.visit_log import VisitLog
from app.schemas.visit_log import (
    format_visit_log_decision_authority_display,
    format_stage_effort_breakdown_display,
)
from app.utils.visit_log_progress_history import format_progress_for_export
from app.utils.visit_log_query_scope import (
    apply_visit_log_role_scope,
    visit_log_standard_loader_options,
)
from app.utils.work_order_query_scope import (
    apply_work_order_role_scope,
    work_order_standard_loader_options,
)
from app.utils.opportunity_query_scope import apply_opportunity_list_role_scope
from app.utils.lead_query_scope import apply_lead_list_role_scope
from app.utils.team_leader_peer_scope import get_peer_team_leader_ids
from app.utils.work_order_pool import work_orders_visible_to_team_leader_filter
from app.models.review import Review
from app.models.group import Group
from app.api.deps import get_current_user, get_current_role
from app.core.permissions import can_view_task
from app.utils.product_utils import convert_product_value_to_label
from app.api.statistics import (
    get_opportunity_converted_amount_statistics,
    get_time_range_statistics,
    get_sales_unit_statistics,
    get_sales_unit_performance_statistics,
    get_requirement_direction_statistics,
    get_member_workloads
)
from app.schemas.statistics import (
    OpportunityConvertedAmountStatistics,
    TimeRangeStatistics,
    SalesUnitStatistics,
    SalesUnitPerformanceStatistics,
    RequirementDirectionGroupStatistics,
    MemberWorkload
)
from datetime import date

router = APIRouter(prefix="/export", tags=["数据导出"])

# 状态映射字典
OPPORTUNITY_STATUS_MAP = {
    "created": "已创建",
    "in_progress": "进行中",
    "lost": "流失",
    "won": "转定",
}

WORK_ORDER_STATUS_MAP = {
    "pending_group_claim": "待组内认领",
    "pending_assign": "待转派",
    "pending_accept": "待接单",
    "accepted": "已接单",
    "in_progress": "已拜访",
    "completed": "已拜访",
    "cancelled": "已取消",
}

TASK_STATUS_MAP = {
    "pending": "待确认",
    "confirmed": "已确认",
    "detail_submitted": "已提交详细需求",
    "dispatched": "已派单",
    "in_progress": "进行中",
    "completed": "已完成",
    "rejected": "已拒绝",
    "cancelled": "已关闭",
}


def get_status_text(status, status_map):
    """获取状态的中文文本
    
    Args:
        status: 枚举状态对象
        status_map: 状态映射字典
    
    Returns:
        str: 状态的中文文本
    """
    # 获取枚举的值（如 "in_progress"）而不是枚举名称
    if hasattr(status, 'value'):
        status_value = status.value
    else:
        status_value = str(status)
    return status_map.get(status_value, str(status))


# 注册中文字体（使用reportlab内置字体或系统字体）
def register_chinese_font():
    """注册中文字体，支持Windows和Linux系统"""
    try:
        # 检查是否已经注册过
        if 'ChineseFont' in pdfmetrics.getRegisteredFontNames():
            logger.info("中文字体已注册: ChineseFont")
            return 'ChineseFont'
        
        font_paths = []
        system = platform.system()
        logger.info(f"检测到系统类型: {system}")
        
        # 获取项目内字体目录路径（Windows和Linux都先检查这个）
        project_font_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'fonts')
        
        if system == 'Windows':
            # 方法1: 优先检查项目内字体目录（如果存在）
            if os.path.exists(project_font_dir):
                logger.info(f"检查项目内字体目录: {project_font_dir}")
                # 优先选择常规字体，避免粗体字体可能的问题
                # 排序：优先选择 .ttc 文件，然后是常规 .ttf（排除 bold/black/heavy 等粗体变体）
                font_files = []
                for file in os.listdir(project_font_dir):
                    if file.lower().endswith(('.ttf', '.ttc', '.otf')):
                        file_lower = file.lower()
                        # 排除粗体变体
                        if any(bold_marker in file_lower for bold_marker in ['bold', 'black', 'heavy', 'b.ttf', 'b.ttc']):
                            continue
                        font_files.append((file, file_lower.endswith('.ttc')))  # (文件名, 是否为ttc)
                
                # 排序：.ttc 优先，然后按文件名排序
                font_files.sort(key=lambda x: (not x[1], x[0]))
                
                for file, _ in font_files:
                    font_path = os.path.join(project_font_dir, file)
                    font_paths.append(font_path)
                    logger.info(f"找到项目内字体: {font_path}")
                    break  # 找到一个就够用了
                
                # 如果没找到常规字体，再尝试所有字体（包括粗体）
                if not font_paths:
                    for file in os.listdir(project_font_dir):
                        if file.lower().endswith(('.ttf', '.ttc', '.otf')):
                            font_path = os.path.join(project_font_dir, file)
                            font_paths.append(font_path)
                            logger.info(f"找到项目内字体（粗体变体）: {font_path}")
                            break
            
            # 方法2: Windows系统字体路径（备用方案）
            if not font_paths:
                windows_font_paths = [
                    "C:/Windows/Fonts/simsun.ttc",  # 宋体
                    "C:/Windows/Fonts/msyh.ttc",    # 微软雅黑
                    "C:/Windows/Fonts/simhei.ttf",  # 黑体
                    "C:/Windows/Fonts/simkai.ttf",  # 楷体
                ]
                font_paths.extend(windows_font_paths)
        elif system == 'Linux':
            # 方法1: 使用fc-list命令查找系统已安装的中文字体（最可靠）
            try:
                result = subprocess.run(
                    ['fc-list', ':lang=zh', 'file'],
                    capture_output=True,
                    text=True,
                    timeout=5
                )
                if result.returncode == 0 and result.stdout.strip():
                    # 解析fc-list输出，提取字体文件路径
                    for line in result.stdout.strip().split('\n'):
                        if line.strip():
                            # fc-list输出格式: /path/to/font.ttf: FontName:style
                            font_path = line.split(':')[0].strip()
                            if os.path.exists(font_path) and font_path.lower().endswith(('.ttf', '.ttc', '.otf')):
                                font_paths.append(font_path)
                                logger.info(f"通过fc-list找到字体: {font_path}")
                                break  # 找到一个就够用了
            except (subprocess.TimeoutExpired, FileNotFoundError, Exception) as e:
                logger.warning(f"fc-list命令执行失败，尝试其他方法: {e}")
            
            # 方法2: 检查项目内字体目录（如果存在）
            if os.path.exists(project_font_dir):
                logger.info(f"检查项目内字体目录: {project_font_dir}")
                for file in os.listdir(project_font_dir):
                    if file.lower().endswith(('.ttf', '.ttc', '.otf')):
                        font_path = os.path.join(project_font_dir, file)
                        font_paths.append(font_path)
                        logger.info(f"找到项目内字体: {font_path}")
                        break
            
            # 方法3: Linux系统常见中文字体路径（备用方案）
            if not font_paths:
                linux_font_dirs = [
                    '/usr/share/fonts/truetype/wqy/',      # 文泉驿字体
                    '/usr/share/fonts/truetype/arphic/',   # 文鼎字体
                    '/usr/share/fonts/truetype/noto/',     # Noto字体
                    '/usr/share/fonts/opentype/noto/',     # Noto OpenType字体
                    '/usr/share/fonts/chinese/',           # 中文字体目录
                    '/usr/share/fonts/truetype/dejavu/',   # DejaVu字体（备用）
                    os.path.expanduser('~/.fonts/'),       # 用户字体目录
                    '/usr/local/share/fonts/',             # 本地字体目录
                ]
                
                # 常见的中文字体文件名模式
                font_patterns = [
                    'wqy-microhei*.ttc',      # 文泉驿微米黑
                    'wqy-zenhei*.ttc',        # 文泉驿正黑
                    'NotoSansCJK*.ttc',       # Noto Sans CJK
                    'NotoSansCJK*.otf',       # Noto Sans CJK (OpenType)
                    'DroidSansFallback*.ttf', # Android字体
                    'AR-PL-UMing-CN*.ttf',    # 文鼎字体
                    'simsun*.ttc',            # 宋体（如果安装了）
                    'msyh*.ttc',              # 微软雅黑（如果安装了）
                    'simhei*.ttf',            # 黑体（如果安装了）
                ]
                
                # 遍历字体目录和模式，查找可用的中文字体
                for font_dir in linux_font_dirs:
                    if os.path.exists(font_dir):
                        for pattern in font_patterns:
                            found_fonts = glob.glob(os.path.join(font_dir, pattern))
                            if found_fonts:
                                font_paths.extend(found_fonts)
                                logger.info(f"在 {font_dir} 找到字体: {found_fonts[0]}")
                                break  # 找到一个就够用了
                        if font_paths:
                            break
                
                # 如果还没找到，尝试在整个字体目录中搜索
                if not font_paths:
                    for font_dir in linux_font_dirs:
                        if os.path.exists(font_dir):
                            # 递归搜索所有字体文件
                            for root, dirs, files in os.walk(font_dir):
                                for file in files:
                                    if file.lower().endswith(('.ttf', '.ttc', '.otf')):
                                        # 检查文件名是否包含中文字体相关关键词
                                        if any(keyword in file.lower() for keyword in 
                                               ['chinese', 'cjk', 'han', 'wqy', 'noto', 'simsun', 'msyh', 'simhei']):
                                            font_path = os.path.join(root, file)
                                            font_paths.append(font_path)
                                            logger.info(f"递归搜索找到字体: {font_path}")
                                            break
                                if font_paths:
                                    break
                            if font_paths:
                                break
        
        # 尝试注册找到的字体
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    # 对于.ttc文件，需要指定字体索引
                    if font_path.endswith('.ttc'):
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path, subfontIndex=0))
                    else:
                        pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                    logger.info(f"✅ 成功注册中文字体: {font_path}")
                    return 'ChineseFont'
                except Exception as e:
                    logger.warning(f"注册字体失败 {font_path}: {e}")
                    # 如果注册失败，继续尝试下一个字体
                    continue
        
        # 如果没有找到系统字体，使用reportlab的内置字体（可能不支持中文，但至少不会报错）
        # 注意：这会导致中文显示为乱码
        logger.error("❌ 未找到可用的中文字体！PDF中的中文将显示为乱码。")
        logger.error("   请参考 backend/docs/LINUX_FONT_SETUP.md 安装中文字体")
        return 'Helvetica'
    except Exception as e:
        # 发生异常时返回默认字体
        logger.error(f"注册中文字体时发生异常: {e}", exc_info=True)
        return 'Helvetica'


def export_tasks_to_excel(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出任务数据到Excel（基于当前激活角色）"""
    # 获取任务列表（根据当前激活角色过滤）
    query = db.query(Task)
    
    if current_role.role == UserRole.TASK_INITIATOR:
        query = query.filter(Task.initiator_id == current_user.id)
    elif current_role.role == UserRole.SALES_CONTACT:
        # 销售单位接口人可以查看销售单位匹配的任务
        # 使用当前激活角色的sales_unit
        user_sales_unit = current_role.sales_unit or current_user.sales_unit
        if user_sales_unit:
            query = query.filter(Task.sales_unit.like(f"%{user_sales_unit}%"))
        else:
            # 兼容旧逻辑
            query = query.filter(Task.sales_contact_id == current_user.id)
    elif current_role.role == UserRole.TEAM_LEADER:
        work_orders = db.query(WorkOrder).filter(
            work_orders_visible_to_team_leader_filter(db, current_user.id)
        ).all()
        task_ids = [wo.task_id for wo in work_orders]
        query = query.filter(Task.id.in_(task_ids))
    elif current_role.role == UserRole.MEMBER:
        work_orders = db.query(WorkOrder).filter(WorkOrder.member_id == current_user.id).all()
        task_ids = [wo.task_id for wo in work_orders]
        query = query.filter(Task.id.in_(task_ids))
    
    tasks = query.order_by(Task.created_at.desc()).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "任务列表"
    
    # 设置表头（客户单位/行业类型在详细需求中维护，任务级字段多为空，不导出；不导出内部任务ID）
    headers = ["任务名称", "销售单位", "开始日期", "结束日期", "FDE人数", "状态", "创建时间"]
    ws.append(headers)
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 填充数据
    for task in tasks:
        status_text = get_status_text(task.status, TASK_STATUS_MAP)
        ws.append([
            task.task_name,
            task.sales_unit,
            task.start_date.strftime("%Y-%m-%d") if task.start_date else "",
            task.end_date.strftime("%Y-%m-%d") if task.end_date else "",
            task.fde_count,
            status_text,
            task.created_at.strftime("%Y-%m-%d %H:%M:%S") if task.created_at else "",
        ])
    
    # 调整列宽
    column_widths = [30, 20, 12, 12, 10, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_work_orders_to_excel(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出工单数据到Excel（基于当前激活角色）"""
    from app.models.user import UserRoleAssociation
    
    query = db.query(WorkOrder).options(*work_order_standard_loader_options())
    query = apply_work_order_role_scope(query, db, current_user, current_role)
    
    work_orders = query.order_by(WorkOrder.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "工单列表"
    
    headers = [
        "工单编号", "任务名称", "客户名称", "客户来源", "发起部门", "组长", "成员", "状态", "拜访时间",
        "客户拜访地址", "客户经理姓名", "客户经理联系方式",
        "接单时间", "完成时间",
    ]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 去除"销售单位 - "前缀的辅助函数
    def remove_sales_unit_prefix(unit: str) -> str:
        if not unit:
            return ""
        # 去除"销售单位-"或"销售单位 - "前缀
        return re.sub(r'^销售单位\s*[-－]\s*', '', unit, flags=re.IGNORECASE).strip()
    
    for wo in work_orders:
        status_text = get_status_text(wo.status, WORK_ORDER_STATUS_MAP)
        # 获取客户名称：优先从详细需求获取，否则从任务获取
        customer_unit = ""
        if wo.detail_requirement:
            customer_unit = wo.detail_requirement.customer_unit or ""
        elif wo.task and wo.task.customer_unit:
            customer_unit = wo.task.customer_unit
        
        # 获取客户来源：从详细需求获取
        customer_source = ""
        if wo.detail_requirement:
            customer_source = wo.detail_requirement.customer_source or ""
        
        # 获取销售单位：优先显示提交人所属的销售单位
        sales_unit = ""
        if wo.detail_requirement and wo.detail_requirement.sales_contact:
            # 获取提交人的销售单位
            sales_role = db.query(UserRoleAssociation).filter(
                UserRoleAssociation.user_id == wo.detail_requirement.sales_contact_id,
                UserRoleAssociation.role == UserRole.SALES_CONTACT,
                UserRoleAssociation.is_active == True
            ).first()
            if sales_role and sales_role.sales_unit:
                sales_unit = remove_sales_unit_prefix(sales_role.sales_unit)
        # 如果没有提交人销售单位，回退到任务的销售单位（兼容旧数据）
        if not sales_unit and wo.task and wo.task.sales_unit:
            sales_unit = remove_sales_unit_prefix(wo.task.sales_unit)
        
        # 获取拜访时间：优先从详细需求获取预期拜访时间
        visit_time = ""
        if wo.detail_requirement and wo.detail_requirement.expected_visit_time:
            visit_time = wo.detail_requirement.expected_visit_time.strftime("%Y-%m-%d %H:%M")
        
        visit_address = ""
        manager_name = ""
        manager_contact = ""
        if wo.detail_requirement:
            visit_address = wo.detail_requirement.customer_visit_address or ""
            manager_name = wo.detail_requirement.customer_manager_name or ""
            manager_contact = wo.detail_requirement.customer_manager_contact or ""
        
        ws.append([
            wo.work_order_no,
            wo.task.task_name if wo.task else "",
            customer_unit,
            customer_source,
            sales_unit,
            wo.team_leader.real_name if wo.team_leader else "",
            wo.member.real_name if wo.member else "",
            status_text,
            visit_time,
            visit_address,
            manager_name,
            manager_contact,
            wo.accepted_at.strftime("%Y-%m-%d %H:%M:%S") if wo.accepted_at else "",
            wo.completed_at.strftime("%Y-%m-%d %H:%M:%S") if wo.completed_at else "",
        ])
    
    column_widths = [20, 30, 25, 20, 25, 15, 15, 12, 20, 28, 14, 18, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_opportunities_to_excel(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出商机数据到Excel（基于当前激活角色）"""
    query = db.query(Opportunity).options(
        joinedload(Opportunity.task),
        joinedload(Opportunity.team_leader)
    )
    query = apply_opportunity_list_role_scope(query, db, current_user, current_role)

    opportunities = query.order_by(Opportunity.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "商机列表"
    
    headers = ["商机编号", "任务名称", "客户单位", "所需产品", "状态", "组长", "转定金额", "流失原因", "创建时间"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for opp in opportunities:
        status_text = get_status_text(opp.status, OPPORTUNITY_STATUS_MAP)
        # 将产品value字符串转换为中文标签
        required_products_label = convert_product_value_to_label(opp.required_products) if opp.required_products else ""
        ws.append([
            opp.opportunity_no,
            opp.task.task_name if opp.task else "",
            opp.customer_unit,
            required_products_label,
            status_text,
            opp.team_leader.real_name if opp.team_leader else "",
            opp.won_amount or "",
            opp.lost_reason or "",
            opp.created_at.strftime("%Y-%m-%d %H:%M:%S") if opp.created_at else "",
        ])
    
    column_widths = [20, 30, 20, 30, 12, 15, 15, 30, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_leads_to_excel(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出线索数据到Excel（基于当前激活角色）"""
    from app.models.lead import Lead
    from app.utils.requirement_direction_utils import convert_requirement_direction_value_to_label

    query = db.query(Lead).options(
        joinedload(Lead.visit_log),
        joinedload(Lead.task),
        joinedload(Lead.member),
        joinedload(Lead.opportunity)
    )
    query = apply_lead_list_role_scope(query, db, current_user, current_role)

    leads = query.order_by(Lead.created_at.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "线索列表"
    
    headers = ["线索ID", "客户名称", "客户需求方向", "详细需求描述", "所属任务", "关联拜访日志ID", "是否已转换为商机", "关联商机ID", "创建人", "创建时间", "更新时间"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for lead in leads:
        # 客户需求方向：与列表 API 一致，将 value/JSON 转为可读标签（避免导出为 raw JSON 字符串）
        requirement_direction_display = ""
        if lead.requirement_direction:
            requirement_direction_display = convert_requirement_direction_value_to_label(db, lead.requirement_direction)
        ws.append([
            lead.id,
            lead.customer_name or "",
            requirement_direction_display,
            lead.detail_description or "",
            lead.task.task_name if lead.task else "",
            lead.visit_log_id or "",
            "是" if lead.opportunity else "否",
            lead.opportunity.id if lead.opportunity else "",
            lead.member.real_name if lead.member else "",
            lead.created_at.strftime("%Y-%m-%d %H:%M:%S") if lead.created_at else "",
            lead.updated_at.strftime("%Y-%m-%d %H:%M:%S") if lead.updated_at else "",
        ])
    
    column_widths = [10, 20, 30, 40, 30, 15, 18, 12, 15, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_visit_logs_to_excel(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出拜访日志数据到Excel（基于当前激活角色）"""
    query = db.query(VisitLog).options(*visit_log_standard_loader_options())
    query = apply_visit_log_role_scope(query, db, current_user, current_role)
    
    visit_logs = query.order_by(VisitLog.visit_date.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "拜访日志列表"
    
    headers = [
        "工单编号", "任务名称", "客户单位", "客户拜访地址", "客户经理", "客户经理联系方式", "组别", "所属销售单位", "行业", "企业类型", "陪跑人员", "创建人", "拜访日期", "拜访对象职位",
        "拜访对象权限", "是否有线索", "线索对应产品", "当前阶段", "阶段人员投入与时长", "推进进展", "推进要求", "是否定开", "定开要求", "预估金额（万元）", "客户是否梳理过需求场景", "需求场景分类", "拜访内容", "备注", "创建时间",
    ]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for vl in visit_logs:
        customer_unit = ""
        if getattr(vl, "customer_unit", None):
            customer_unit = vl.customer_unit or ""
        elif vl.work_order and vl.work_order.detail_requirement:
            customer_unit = vl.work_order.detail_requirement.customer_unit or ""
        elif vl.work_order and vl.work_order.task and vl.work_order.task.customer_unit:
            customer_unit = vl.work_order.task.customer_unit or ""
        visit_addr = ""
        if getattr(vl, "customer_visit_address", None):
            visit_addr = vl.customer_visit_address or ""
        elif vl.work_order and vl.work_order.detail_requirement:
            visit_addr = vl.work_order.detail_requirement.customer_visit_address or ""
        mgr_name = ""
        if getattr(vl, "customer_manager_name", None):
            mgr_name = vl.customer_manager_name or ""
        elif vl.work_order and vl.work_order.detail_requirement:
            mgr_name = vl.work_order.detail_requirement.customer_manager_name or ""
        mgr_contact = ""
        if getattr(vl, "customer_manager_contact", None):
            mgr_contact = vl.customer_manager_contact or ""
        elif vl.work_order and vl.work_order.detail_requirement:
            mgr_contact = vl.work_order.detail_requirement.customer_manager_contact or ""
        sales_unit_cell = ""
        if getattr(vl, "sales_unit", None):
            sales_unit_cell = vl.sales_unit or ""
        elif vl.work_order and vl.work_order.task:
            sales_unit_cell = vl.work_order.task.sales_unit or ""
        requirement_scenario_category_text = ""
        if getattr(vl, "requirement_scenario_category", None):
            from app.utils.visit_log_requirement_scenario_utils import (
                format_visit_log_requirement_scenario_category_for_export,
            )

            requirement_scenario_category_text = (
                format_visit_log_requirement_scenario_category_for_export(
                    db, vl.requirement_scenario_category
                )
            )
        clue_related_products_text = ""
        if getattr(vl, "clue_related_products", None):
            from app.utils.product_utils import convert_product_value_to_label

            clue_related_products_text = convert_product_value_to_label(
                vl.clue_related_products, db
            )
        ws.append([
            vl.work_order.work_order_no if vl.work_order else "",
            vl.work_order.task.task_name if vl.work_order and vl.work_order.task else "",
            customer_unit,
            visit_addr,
            mgr_name,
            mgr_contact,
            vl.group_name or "",
            sales_unit_cell,
            getattr(vl, "industry", None) or "",
            getattr(vl, "enterprise_type", None) or "",
            getattr(vl, "escort_staff", None) or "",
            vl.member.real_name if vl.member else "",
            vl.visit_date.strftime("%Y-%m-%d") if vl.visit_date else "",
            vl.visit_object_position or "",
            format_visit_log_decision_authority_display(vl.has_decision_authority),
            "是" if getattr(vl, "has_clue", False) else "否",
            clue_related_products_text,
            getattr(vl, "current_stage", None) or "",
            format_stage_effort_breakdown_display(
                getattr(vl, "stage_effort_breakdown", None)
            ),
            format_progress_for_export(
                getattr(vl, "promotion_progress_history", None),
                getattr(vl, "promotion_progress", None),
            ),
            getattr(vl, "promotion_requirements", None) or "",
            "是" if getattr(vl, "is_customized_development", False) else "否",
            getattr(vl, "customized_development_requirements", None) or "",
            getattr(vl, "project_amount", None) or "",
            "是" if getattr(vl, "has_requirement_scenario_sorted", False) else "否",
            requirement_scenario_category_text,
            vl.visit_content or "",
            getattr(vl, "remark", None) or "",
            vl.created_at.strftime("%Y-%m-%d %H:%M:%S") if vl.created_at else "",
        ])
    
    column_widths = [
        20, 28, 18, 24, 12, 14, 14, 18, 14, 12, 14, 15, 12, 20, 18, 12, 28, 16, 36,
        24, 24, 12, 18, 28, 32, 16, 24, 20, 20,
    ]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/tasks/excel")
def export_tasks_excel(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出任务数据为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_tasks_to_excel(db, current_user, current_role)
    filename = f"任务列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/work-orders/excel")
def export_work_orders_excel(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出工单数据为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_work_orders_to_excel(db, current_user, current_role)
    filename = f"工单列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/opportunities/excel")
def export_opportunities_excel(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出商机数据为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_opportunities_to_excel(db, current_user, current_role)
    filename = f"商机列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/leads/excel")
def export_leads_excel(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出线索数据为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_leads_to_excel(db, current_user, current_role)
    filename = f"线索列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/visit-logs/excel")
def export_visit_logs_excel(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出拜访日志数据为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_visit_logs_to_excel(db, current_user, current_role)
    filename = f"拜访日志列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


# ==================== PDF导出功能 ====================

def export_tasks_to_pdf(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出任务数据到PDF（基于当前激活角色）"""
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    # 获取任务列表（根据当前激活角色过滤）
    query = db.query(Task)
    
    if current_role.role == UserRole.TASK_INITIATOR:
        query = query.filter(Task.initiator_id == current_user.id)
    elif current_role.role == UserRole.SALES_CONTACT:
        # 销售单位接口人可以查看销售单位匹配的任务
        # 使用当前激活角色的sales_unit
        user_sales_unit = current_role.sales_unit or current_user.sales_unit
        if user_sales_unit:
            query = query.filter(Task.sales_unit.like(f"%{user_sales_unit}%"))
        else:
            # 兼容旧逻辑
            query = query.filter(Task.sales_contact_id == current_user.id)
    elif current_role.role == UserRole.TEAM_LEADER:
        work_orders = db.query(WorkOrder).filter(
            work_orders_visible_to_team_leader_filter(db, current_user.id)
        ).all()
        task_ids = [wo.task_id for wo in work_orders]
        query = query.filter(Task.id.in_(task_ids))
    elif current_role.role == UserRole.MEMBER:
        work_orders = db.query(WorkOrder).filter(WorkOrder.member_id == current_user.id).all()
        task_ids = [wo.task_id for wo in work_orders]
        query = query.filter(Task.id.in_(task_ids))
    
    tasks = query.order_by(Task.created_at.desc()).all()
    
    # 创建PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    # 标题
    elements.append(Paragraph("任务列表", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 生成时间
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 表格数据（不导出内部任务ID；客户单位/行业类型见详细需求）
    data = [["任务名称", "销售单位", "开始日期", "结束日期", "FDE人数", "状态"]]
    
    for task in tasks:
        status_text = get_status_text(task.status, TASK_STATUS_MAP)
        data.append([
            task.task_name[:20] if len(task.task_name) > 20 else task.task_name,
            task.sales_unit[:15] if len(task.sales_unit) > 15 else task.sales_unit,
            task.start_date.strftime("%Y-%m-%d") if task.start_date else "",
            task.end_date.strftime("%Y-%m-%d") if task.end_date else "",
            str(task.fde_count),
            status_text,
        ])
    
    # 创建表格
    table = Table(data, colWidths=[2.2*inch, 1.4*inch, 1*inch, 1*inch, 0.8*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(tasks)} 条记录", normal_style))
    
    # 构建PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_work_orders_to_pdf(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出工单数据到PDF（基于当前激活角色）"""
    from app.models.user import UserRoleAssociation
    
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    query = db.query(WorkOrder).options(*work_order_standard_loader_options())
    query = apply_work_order_role_scope(query, db, current_user, current_role)
    
    work_orders = query.order_by(WorkOrder.created_at.desc()).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("工单列表", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    data = [[
        "工单编号", "任务名称", "客户名称", "客户来源", "发起部门", "组长", "成员", "状态", "拜访时间",
        "客户拜访地址", "客户经理姓名", "客户经理联系方式",
    ]]
    
    # 去除"销售单位 - "前缀的辅助函数
    def remove_sales_unit_prefix(unit: str) -> str:
        if not unit:
            return ""
        # 去除"销售单位-"或"销售单位 - "前缀
        return re.sub(r'^销售单位\s*[-－]\s*', '', unit, flags=re.IGNORECASE).strip()
    
    for wo in work_orders:
        status_text = get_status_text(wo.status, WORK_ORDER_STATUS_MAP)
        # 获取客户名称：优先从详细需求获取，否则从任务获取
        customer_unit = ""
        if wo.detail_requirement:
            customer_unit = (wo.detail_requirement.customer_unit or "")[:15]
        elif wo.task and wo.task.customer_unit:
            customer_unit = wo.task.customer_unit[:15]
        
        # 获取客户来源：从详细需求获取
        customer_source = ""
        if wo.detail_requirement:
            customer_source = (wo.detail_requirement.customer_source or "")[:20]
        
        # 获取销售单位：优先显示提交人所属的销售单位
        sales_unit = ""
        if wo.detail_requirement and wo.detail_requirement.sales_contact:
            # 获取提交人的销售单位
            sales_role = db.query(UserRoleAssociation).filter(
                UserRoleAssociation.user_id == wo.detail_requirement.sales_contact_id,
                UserRoleAssociation.role == UserRole.SALES_CONTACT,
                UserRoleAssociation.is_active == True
            ).first()
            if sales_role and sales_role.sales_unit:
                sales_unit = remove_sales_unit_prefix(sales_role.sales_unit)[:20]
        # 如果没有提交人销售单位，回退到任务的销售单位（兼容旧数据）
        if not sales_unit and wo.task and wo.task.sales_unit:
            sales_unit = remove_sales_unit_prefix(wo.task.sales_unit)[:20]
        
        # 获取拜访时间：优先从详细需求获取预期拜访时间
        visit_time = ""
        if wo.detail_requirement and wo.detail_requirement.expected_visit_time:
            visit_time = wo.detail_requirement.expected_visit_time.strftime("%Y-%m-%d %H:%M")
        
        visit_address = ""
        manager_name = ""
        manager_contact = ""
        if wo.detail_requirement:
            raw_addr = wo.detail_requirement.customer_visit_address or ""
            visit_address = raw_addr[:22] + "…" if len(raw_addr) > 22 else raw_addr
            raw_mn = wo.detail_requirement.customer_manager_name or ""
            manager_name = raw_mn[:10] + "…" if len(raw_mn) > 10 else raw_mn
            raw_mc = wo.detail_requirement.customer_manager_contact or ""
            manager_contact = raw_mc[:16] + "…" if len(raw_mc) > 16 else raw_mc
        
        data.append([
            wo.work_order_no[:15] if len(wo.work_order_no) > 15 else wo.work_order_no,
            (wo.task.task_name[:15] if wo.task and len(wo.task.task_name) > 15 else (wo.task.task_name if wo.task else "")),
            customer_unit,
            customer_source,
            sales_unit,
            wo.team_leader.real_name[:10] if wo.team_leader and wo.team_leader.real_name else "",
            wo.member.real_name[:10] if wo.member and wo.member.real_name else "",
            status_text,
            visit_time,
            visit_address,
            manager_name,
            manager_contact,
        ])
    
    table = Table(
        data,
        colWidths=[
            0.52 * inch, 0.62 * inch, 0.48 * inch, 0.5 * inch, 0.48 * inch,
            0.42 * inch, 0.42 * inch, 0.42 * inch, 0.55 * inch,
            0.62 * inch, 0.45 * inch, 0.52 * inch,
        ],
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(work_orders)} 条记录", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_opportunities_to_pdf(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出商机数据到PDF（基于当前激活角色）"""
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    query = db.query(Opportunity).options(
        joinedload(Opportunity.task),
        joinedload(Opportunity.team_leader)
    )
    query = apply_opportunity_list_role_scope(query, db, current_user, current_role)

    opportunities = query.order_by(Opportunity.created_at.desc()).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("商机列表", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    data = [["商机编号", "任务名称", "客户单位", "所需产品", "状态", "转定金额"]]
    
    for opp in opportunities:
        status_text = get_status_text(opp.status, OPPORTUNITY_STATUS_MAP)
        # 将产品value字符串转换为中文标签
        required_products_label = convert_product_value_to_label(opp.required_products) if opp.required_products else ""
        data.append([
            opp.opportunity_no[:15] if len(opp.opportunity_no) > 15 else opp.opportunity_no,
            (opp.task.task_name[:20] if opp.task and len(opp.task.task_name) > 20 else (opp.task.task_name if opp.task else "")),
            opp.customer_unit[:15] if len(opp.customer_unit) > 15 else opp.customer_unit,
            (required_products_label[:20] if required_products_label and len(required_products_label) > 20 else required_products_label),
            status_text,
            str(opp.won_amount) if opp.won_amount else "",
        ])
    
    table = Table(data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.5*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(opportunities)} 条记录", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_leads_to_pdf(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出线索数据到PDF（基于当前激活角色）"""
    from app.models.lead import Lead
    from app.utils.requirement_direction_utils import convert_requirement_direction_value_to_label

    # 注册中文字体
    chinese_font = register_chinese_font()
    
    query = db.query(Lead).options(
        joinedload(Lead.visit_log),
        joinedload(Lead.task),
        joinedload(Lead.member),
        joinedload(Lead.opportunity)
    )
    query = apply_lead_list_role_scope(query, db, current_user, current_role)

    leads = query.order_by(Lead.created_at.desc()).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("线索列表", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    data = [["线索ID", "客户名称", "需求方向", "所属任务", "已转商机", "创建人"]]

    for lead in leads:
        requirement_direction_display = ""
        if lead.requirement_direction:
            requirement_direction_display = convert_requirement_direction_value_to_label(db, lead.requirement_direction)
        if len(requirement_direction_display) > 20:
            requirement_direction_display = requirement_direction_display[:20] + "…"
        data.append([
            str(lead.id),
            (lead.customer_name[:15] if lead.customer_name and len(lead.customer_name) > 15 else (lead.customer_name or "")),
            requirement_direction_display,
            (lead.task.task_name[:20] if lead.task and lead.task.task_name and len(lead.task.task_name) > 20 else (lead.task.task_name if lead.task else "")),
            "是" if lead.opportunity else "否",
            (lead.member.real_name[:10] if lead.member and lead.member.real_name else ""),
        ])
    
    table = Table(data, colWidths=[0.8*inch, 1.5*inch, 2*inch, 2*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(leads)} 条记录", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_visit_logs_to_pdf(db: Session, current_user: User, current_role) -> io.BytesIO:
    """导出拜访日志数据到PDF（基于当前激活角色）"""
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    query = db.query(VisitLog).options(*visit_log_standard_loader_options())
    query = apply_visit_log_role_scope(query, db, current_user, current_role)
    
    visit_logs = query.order_by(VisitLog.visit_date.desc()).all()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("拜访日志列表", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    data = [[
        "工单编号", "任务名称", "客户单位", "拜访地址", "客户经理", "联系方式", "组别", "所属销售单位", "行业", "企业类型", "陪跑人员", "创建人", "拜访日期", "拜访对象职位",
        "拜访对象权限", "是否有线索", "线索对应产品", "当前阶段", "阶段人员与时长", "推进进展", "推进要求", "是否定开", "定开要求", "预估金额（万元）", "客户是否梳理过需求场景", "需求场景分类", "拜访内容", "备注",
    ]]
    
    for vl in visit_logs:
        raw_cu = ""
        if getattr(vl, "customer_unit", None):
            raw_cu = vl.customer_unit or ""
        elif vl.work_order and vl.work_order.detail_requirement:
            raw_cu = vl.work_order.detail_requirement.customer_unit or ""
        elif vl.work_order and vl.work_order.task and vl.work_order.task.customer_unit:
            raw_cu = vl.work_order.task.customer_unit or ""
        customer_unit = (raw_cu[:12] + "…" if len(raw_cu) > 12 else raw_cu) if raw_cu else ""
        raw_va = ""
        if getattr(vl, "customer_visit_address", None):
            raw_va = vl.customer_visit_address or ""
        elif vl.work_order and vl.work_order.detail_requirement:
            raw_va = vl.work_order.detail_requirement.customer_visit_address or ""
        va_cell = raw_va[:10] + ("…" if len(raw_va) > 10 else "") if raw_va else ""
        raw_mn = ""
        if getattr(vl, "customer_manager_name", None):
            raw_mn = vl.customer_manager_name or ""
        elif vl.work_order and vl.work_order.detail_requirement:
            raw_mn = vl.work_order.detail_requirement.customer_manager_name or ""
        mn_cell = raw_mn[:8] + ("…" if len(raw_mn) > 8 else "") if raw_mn else ""
        raw_mc = ""
        if getattr(vl, "customer_manager_contact", None):
            raw_mc = vl.customer_manager_contact or ""
        elif vl.work_order and vl.work_order.detail_requirement:
            raw_mc = vl.work_order.detail_requirement.customer_manager_contact or ""
        mc_cell = raw_mc[:10] + ("…" if len(raw_mc) > 10 else "") if raw_mc else ""
        gn = vl.group_name or ""
        group_name_cell = gn[:8] + ("…" if len(gn) > 8 else "") if gn else ""
        raw_su = ""
        if getattr(vl, "sales_unit", None):
            raw_su = vl.sales_unit or ""
        elif vl.work_order and vl.work_order.task:
            raw_su = vl.work_order.task.sales_unit or ""
        sales_unit_cell = raw_su[:10] + ("…" if len(raw_su) > 10 else "") if raw_su else ""
        raw_ind = getattr(vl, "industry", None) or ""
        ind = raw_ind[:8] + ("…" if len(raw_ind) > 8 else "")
        et = getattr(vl, "enterprise_type", None) or ""
        esc = getattr(vl, "escort_staff", None) or ""
        esc_cell = esc[:8] + ("…" if len(esc) > 8 else "") if esc else ""
        visit_content = (vl.visit_content[:28] if vl.visit_content and len(vl.visit_content) > 28 else (vl.visit_content or ""))
        raw_rm = getattr(vl, "remark", None) or ""
        remark_cell = raw_rm[:20] + ("…" if len(raw_rm) > 20 else "") if raw_rm else ""
        requirement_scenario_category_cell = ""
        if getattr(vl, "requirement_scenario_category", None):
            from app.utils.visit_log_requirement_scenario_utils import (
                format_visit_log_requirement_scenario_category_for_export,
            )

            raw_rsc = format_visit_log_requirement_scenario_category_for_export(
                db, vl.requirement_scenario_category
            )
            requirement_scenario_category_cell = (
                raw_rsc[:20] + ("…" if len(raw_rsc) > 20 else "")
            ) if raw_rsc else ""
        clue_related_products_cell = ""
        if getattr(vl, "clue_related_products", None):
            from app.utils.product_utils import convert_product_value_to_label

            raw_crp = convert_product_value_to_label(vl.clue_related_products, db)
            clue_related_products_cell = (
                raw_crp[:20] + ("…" if len(raw_crp) > 20 else "")
            ) if raw_crp else ""
        raw_pp = format_progress_for_export(
            getattr(vl, "promotion_progress_history", None),
            getattr(vl, "promotion_progress", None),
        )
        promotion_progress_cell = raw_pp[:16] + ("…" if len(raw_pp) > 16 else "") if raw_pp else ""
        raw_pr = getattr(vl, "promotion_requirements", None) or ""
        promotion_requirements_cell = raw_pr[:16] + ("…" if len(raw_pr) > 16 else "") if raw_pr else ""
        raw_cdr = getattr(vl, "customized_development_requirements", None) or ""
        cdr_cell = raw_cdr[:14] + ("…" if len(raw_cdr) > 14 else "") if raw_cdr else ""
        raw_stage_eff = format_stage_effort_breakdown_display(
            getattr(vl, "stage_effort_breakdown", None)
        )
        stage_effort_cell = (
            raw_stage_eff[:24] + ("…" if len(raw_stage_eff) > 24 else "")
        ) if raw_stage_eff else ""
        data.append([
            vl.work_order.work_order_no[:12] if vl.work_order and len(vl.work_order.work_order_no) > 12 else (vl.work_order.work_order_no if vl.work_order else ""),
            (vl.work_order.task.task_name[:16] if vl.work_order and vl.work_order.task and len(vl.work_order.task.task_name) > 16 else (vl.work_order.task.task_name if vl.work_order and vl.work_order.task else "")),
            customer_unit,
            va_cell,
            mn_cell,
            mc_cell,
            group_name_cell,
            sales_unit_cell,
            ind,
            et[:6] + ("…" if len(et) > 6 else "") if et else "",
            esc_cell,
            vl.member.real_name[:8] if vl.member and vl.member.real_name else "",
            vl.visit_date.strftime("%Y-%m-%d") if vl.visit_date else "",
            (vl.visit_object_position[:12] if vl.visit_object_position and len(vl.visit_object_position) > 12 else (vl.visit_object_position or "")),
            format_visit_log_decision_authority_display(vl.has_decision_authority),
            "是" if getattr(vl, "has_clue", False) else "否",
            clue_related_products_cell,
            (getattr(vl, "current_stage", None) or "")[:10],
            stage_effort_cell,
            promotion_progress_cell,
            promotion_requirements_cell,
            "是" if getattr(vl, "is_customized_development", False) else "否",
            cdr_cell,
            (getattr(vl, "project_amount", None) or "")[:12],
            "是" if getattr(vl, "has_requirement_scenario_sorted", False) else "否",
            requirement_scenario_category_cell,
            visit_content,
            remark_cell,
        ])
    
    _vl_pdf_col_widths = [
        0.52 * inch, 0.68 * inch, 0.48 * inch, 0.48 * inch, 0.42 * inch, 0.48 * inch,
        0.38 * inch, 0.38 * inch, 0.38 * inch, 0.38 * inch, 0.4 * inch, 0.38 * inch,
        0.45 * inch, 0.5 * inch,
        0.32 * inch, 0.35 * inch, 0.35 * inch, 0.38 * inch, 0.55 * inch, 0.42 * inch,
        0.45 * inch, 0.32 * inch, 0.5 * inch, 0.36 * inch, 0.32 * inch, 0.38 * inch,
        0.5 * inch, 0.38 * inch,
    ]
    table = Table(data, colWidths=_vl_pdf_col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(visit_logs)} 条记录", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/tasks/pdf")
def export_tasks_pdf(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出任务数据为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_tasks_to_pdf(db, current_user, current_role)
    filename = f"任务列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/work-orders/pdf")
def export_work_orders_pdf(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出工单数据为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_work_orders_to_pdf(db, current_user, current_role)
    filename = f"工单列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/opportunities/pdf")
def export_opportunities_pdf(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出商机数据为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_opportunities_to_pdf(db, current_user, current_role)
    filename = f"商机列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/leads/pdf")
def export_leads_pdf(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出线索数据为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_leads_to_pdf(db, current_user, current_role)
    filename = f"线索列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/visit-logs/pdf")
def export_visit_logs_pdf(
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出拜访日志数据为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    output = export_visit_logs_to_pdf(db, current_user, current_role)
    filename = f"拜访日志列表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


# ==================== 转订商机总金额统计导出功能 ====================

def export_opportunity_converted_amount_to_excel(
    db: Session, 
    current_user: User, 
    current_role,
    group_id: Optional[int] = None,
    include_member_details: bool = False,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出转订商机总金额统计为Excel（基于当前激活角色）"""
    from app.api.statistics import get_opportunity_converted_amount_statistics
    
    # 获取统计数据
    statistics = get_opportunity_converted_amount_statistics(
        db, current_user, current_role, group_id, include_member_details, start_date, end_date
    )
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "转订商机总金额统计"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据样式
    data_alignment = Alignment(horizontal="center", vertical="center")
    total_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
    
    # 写入标题
    ws['A1'] = "序号"
    ws['B1'] = "名称"
    ws['C1'] = "转订商机数"
    ws['D1'] = "转订总金额（元）"
    ws['E1'] = "平均金额（元）"
    
    # 设置标题样式
    for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = header_alignment
    
    # 写入数据
    row = 2
    total_converted_count = 0
    total_amount = 0.0
    
    for idx, stat in enumerate(statistics, start=1):
        ws[f'A{row}'] = idx
        ws[f'A{row}'].alignment = data_alignment
        
        # 名称列
        display_name = stat.member_name or stat.group_name or "未知"
        ws[f'B{row}'] = display_name
        ws[f'B{row}'].alignment = data_alignment
        
        # 转订商机数
        ws[f'C{row}'] = stat.converted_count
        ws[f'C{row}'].alignment = data_alignment
        
        # 转订总金额
        ws[f'D{row}'] = stat.total_amount
        ws[f'D{row}'].alignment = data_alignment
        ws[f'D{row}'].number_format = '#,##0.00'
        
        # 平均金额
        avg_amount = (stat.total_amount / stat.converted_count) if stat.converted_count > 0 else 0
        ws[f'E{row}'] = avg_amount
        ws[f'E{row}'].alignment = data_alignment
        ws[f'E{row}'].number_format = '#,##0.00'
        
        # 总计行特殊样式
        is_total = display_name in ["组总计", "总计"]
        if is_total:
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].fill = total_fill
                ws[f'{col}{row}'].font = Font(bold=True)
        
        total_converted_count += stat.converted_count
        total_amount += stat.total_amount
        row += 1
    
    # 如果有汇总数据，添加汇总行
    if len(statistics) > 1:
        ws[f'A{row}'] = "合计"
        ws[f'B{row}'] = f"共{len(statistics)}项"
        ws[f'C{row}'] = total_converted_count
        ws[f'D{row}'] = total_amount
        ws[f'E{row}'] = (total_amount / total_converted_count) if total_converted_count > 0 else 0
        
        # 设置汇总行样式
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].alignment = data_alignment
        
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'E{row}'].number_format = '#,##0.00'
    
    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_opportunity_converted_amount_to_pdf(
    db: Session, 
    current_user: User, 
    current_role,
    group_id: Optional[int] = None,
    include_member_details: bool = False,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出转订商机总金额统计为PDF（基于当前激活角色）"""
    from app.api.statistics import get_opportunity_converted_amount_statistics
    
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    # 获取统计数据
    statistics = get_opportunity_converted_amount_statistics(
        db, current_user, current_role, group_id, include_member_details, start_date, end_date
    )
    
    # 创建PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("转订商机总金额统计", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 添加时间范围信息
    date_range_text = "全部时间"
    if start_date or end_date:
        date_range_text = f"{start_date.strftime('%Y-%m-%d') if start_date else ''} 至 {end_date.strftime('%Y-%m-%d') if end_date else ''}"
    elements.append(Paragraph(f"时间范围：{date_range_text}", normal_style))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 准备表格数据
    data = [["序号", "名称", "转订商机数", "转订总金额（元）", "平均金额（元）"]]
    
    total_converted_count = 0
    total_amount = 0.0
    
    for idx, stat in enumerate(statistics, start=1):
        display_name = stat.member_name or stat.group_name or "未知"
        avg_amount = (stat.total_amount / stat.converted_count) if stat.converted_count > 0 else 0
        
        data.append([
            str(idx),
            display_name[:20] if len(display_name) > 20 else display_name,
            str(stat.converted_count),
            f"{stat.total_amount:,.2f}",
            f"{avg_amount:,.2f}"
        ])
        
        total_converted_count += stat.converted_count
        total_amount += stat.total_amount
    
    # 如果有汇总数据，添加汇总行
    if len(statistics) > 1:
        avg_total = (total_amount / total_converted_count) if total_converted_count > 0 else 0
        data.append([
            "",
            f"合计（共{len(statistics)}项）",
            str(total_converted_count),
            f"{total_amount:,.2f}",
            f"{avg_total:,.2f}"
        ])
    
    # 创建表格
    table = Table(data, colWidths=[0.8*inch, 1.5*inch, 1*inch, 1.2*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C5E0B4')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        ('FONTNAME', (0, -1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(statistics)} 项统计记录", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/statistics/opportunity-converted-amount/excel")
def export_opportunity_converted_amount_excel(
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    include_member_details: bool = Query(False, description="是否包含成员明细（仅组长可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出转订商机总金额统计为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    if include_member_details and current_role.role != UserRole.TEAM_LEADER:
        raise HTTPException(status_code=403, detail="只有组长可以查看成员明细")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_opportunity_converted_amount_to_excel(
        db, current_user, current_role, group_id, include_member_details, start_date_obj, end_date_obj
    )
    filename = f"转订商机总金额统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/statistics/opportunity-converted-amount/pdf")
def export_opportunity_converted_amount_pdf(
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    include_member_details: bool = Query(False, description="是否包含成员明细（仅组长可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出转订商机总金额统计为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    if include_member_details and current_role.role != UserRole.TEAM_LEADER:
        raise HTTPException(status_code=403, detail="只有组长可以查看成员明细")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_opportunity_converted_amount_to_pdf(
        db, current_user, current_role, group_id, include_member_details,
        start_date_obj, end_date_obj
    )
    filename = f"转订商机总金额统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


# ==================== 时间段趋势统计导出功能 ====================

def export_time_range_statistics_to_excel(
    db: Session, 
    current_user: User, 
    current_role,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    group_by: str = "day",
    group_id: Optional[int] = None
) -> io.BytesIO:
    """导出时间段趋势统计为Excel（基于当前激活角色）"""
    # 获取统计数据
    statistics = get_time_range_statistics(
        db, current_user, current_role, start_date, end_date, group_by, group_id
    )
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "时间段趋势统计"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据样式
    data_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入标题
    ws['A1'] = "日期"
    ws['B1'] = "任务数"
    ws['C1'] = "工单数"
    ws['D1'] = "拜访日志数"
    ws['E1'] = "商机数"
    
    # 设置标题样式
    for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = header_alignment
    
    # 写入数据
    row = 2
    total_task_count = 0
    total_work_order_count = 0
    total_visit_log_count = 0
    total_opportunity_count = 0
    
    for stat in statistics:
        # 日期
        date_str = stat.date.strftime('%Y-%m-%d')
        ws[f'A{row}'] = date_str
        ws[f'A{row}'].alignment = data_alignment
        
        # 任务数
        ws[f'B{row}'] = stat.task_count
        ws[f'B{row}'].alignment = data_alignment
        
        # 工单数
        ws[f'C{row}'] = stat.work_order_count
        ws[f'C{row}'].alignment = data_alignment
        
        # 拜访日志数
        ws[f'D{row}'] = stat.visit_log_count
        ws[f'D{row}'].alignment = data_alignment
        
        # 商机数
        ws[f'E{row}'] = stat.opportunity_count
        ws[f'E{row}'].alignment = data_alignment
        
        total_task_count += stat.task_count
        total_work_order_count += stat.work_order_count
        total_visit_log_count += stat.visit_log_count
        total_opportunity_count += stat.opportunity_count
        row += 1
    
    # 添加汇总行
    if len(statistics) > 0:
        total_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        ws[f'A{row}'] = "合计"
        ws[f'B{row}'] = total_task_count
        ws[f'C{row}'] = total_work_order_count
        ws[f'D{row}'] = total_visit_log_count
        ws[f'E{row}'] = total_opportunity_count
        
        # 设置汇总行样式
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].alignment = data_alignment
    
    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_time_range_statistics_to_pdf(
    db: Session, 
    current_user: User, 
    current_role,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    group_by: str = "day",
    group_id: Optional[int] = None
) -> io.BytesIO:
    """导出时间段趋势统计为PDF（基于当前激活角色）"""
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    # 获取统计数据
    statistics = get_time_range_statistics(
        db, current_user, current_role, start_date, end_date, group_by, group_id
    )
    
    # 创建PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("时间段趋势统计", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 添加时间范围信息
    date_range_text = "全部时间"
    if start_date or end_date:
        date_range_text = f"{start_date.strftime('%Y-%m-%d') if start_date else ''} 至 {end_date.strftime('%Y-%m-%d') if end_date else ''}"
    elements.append(Paragraph(f"时间范围：{date_range_text}", normal_style))
    elements.append(Paragraph(f"分组方式：{group_by}", normal_style))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 准备表格数据
    data = [["日期", "任务数", "工单数", "拜访日志数", "商机数"]]
    
    total_task_count = 0
    total_work_order_count = 0
    total_visit_log_count = 0
    total_opportunity_count = 0
    
    for stat in statistics:
        date_str = stat.date.strftime('%Y-%m-%d')
        data.append([
            date_str,
            str(stat.task_count),
            str(stat.work_order_count),
            str(stat.visit_log_count),
            str(stat.opportunity_count)
        ])
        
        total_task_count += stat.task_count
        total_work_order_count += stat.work_order_count
        total_visit_log_count += stat.visit_log_count
        total_opportunity_count += stat.opportunity_count
    
    # 添加汇总行
    if len(statistics) > 0:
        data.append([
            "合计",
            str(total_task_count),
            str(total_work_order_count),
            str(total_visit_log_count),
            str(total_opportunity_count)
        ])
    
    # 创建表格
    table = Table(data, colWidths=[1.2*inch, 1*inch, 1*inch, 1.2*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C5E0B4')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        ('FONTNAME', (0, -1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(statistics)} 条统计记录", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/statistics/time-range/excel")
def export_time_range_statistics_excel(
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    group_by: str = Query("day", description="分组方式：day/week/month"),
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出时间段趋势统计为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    
    if group_by not in ["day", "week", "month"]:
        raise HTTPException(status_code=400, detail="group_by 必须是 day、week 或 month")
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_time_range_statistics_to_excel(
        db, current_user, current_role, start_date_obj, end_date_obj, group_by, group_id
    )
    filename = f"时间段趋势统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/statistics/time-range/pdf")
def export_time_range_statistics_pdf(
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    group_by: str = Query("day", description="分组方式：day/week/month"),
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出时间段趋势统计为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    
    if group_by not in ["day", "week", "month"]:
        raise HTTPException(status_code=400, detail="group_by 必须是 day、week 或 month")
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_time_range_statistics_to_pdf(
        db, current_user, current_role, start_date_obj, end_date_obj, group_by, group_id
    )
    filename = f"时间段趋势统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


# ==================== 部门数据统计导出功能 ====================

def export_sales_unit_statistics_to_excel(
    db: Session, 
    current_user: User, 
    current_role,
    group_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出部门数据统计为Excel（基于当前激活角色）"""
    # 获取统计数据
    statistics = get_sales_unit_statistics(
        db, current_user, current_role, group_id, start_date, end_date
    )
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "部门数据统计"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据样式
    data_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入标题
    ws['A1'] = "销售单位"
    ws['B1'] = "任务数"
    ws['C1'] = "工单数"
    ws['D1'] = "商机数"
    ws['E1'] = "转化率（%）"
    
    # 设置标题样式
    for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = header_alignment
    
    # 写入数据
    row = 2
    total_task_count = 0
    total_work_order_count = 0
    total_opportunity_count = 0
    
    for stat in statistics:
        # 销售单位
        ws[f'A{row}'] = stat.sales_unit
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
        
        # 任务数
        ws[f'B{row}'] = stat.task_count
        ws[f'B{row}'].alignment = data_alignment
        
        # 工单数
        ws[f'C{row}'] = stat.work_order_count
        ws[f'C{row}'].alignment = data_alignment
        
        # 商机数
        ws[f'D{row}'] = stat.opportunity_count
        ws[f'D{row}'].alignment = data_alignment
        
        # 转化率
        ws[f'E{row}'] = stat.conversion_rate
        ws[f'E{row}'].alignment = data_alignment
        ws[f'E{row}'].number_format = '0.00'
        
        total_task_count += stat.task_count
        total_work_order_count += stat.work_order_count
        total_opportunity_count += stat.opportunity_count
        row += 1
    
    # 添加汇总行
    if len(statistics) > 0:
        total_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        total_conversion_rate = (total_opportunity_count / total_task_count * 100) if total_task_count > 0 else 0.0
        
        ws[f'A{row}'] = "合计"
        ws[f'B{row}'] = total_task_count
        ws[f'C{row}'] = total_work_order_count
        ws[f'D{row}'] = total_opportunity_count
        ws[f'E{row}'] = round(total_conversion_rate, 2)
        
        # 设置汇总行样式
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].alignment = data_alignment
        
        ws[f'E{row}'].number_format = '0.00'
    
    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_sales_unit_statistics_to_pdf(
    db: Session, 
    current_user: User, 
    current_role,
    group_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出部门数据统计为PDF（基于当前激活角色）"""
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    # 获取统计数据
    statistics = get_sales_unit_statistics(
        db, current_user, current_role, group_id, start_date, end_date
    )
    
    # 创建PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("部门数据统计", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 添加时间范围信息
    date_range_text = "全部时间"
    if start_date or end_date:
        date_range_text = f"{start_date.strftime('%Y-%m-%d') if start_date else ''} 至 {end_date.strftime('%Y-%m-%d') if end_date else ''}"
    elements.append(Paragraph(f"时间范围：{date_range_text}", normal_style))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 准备表格数据
    data = [["销售单位", "任务数", "工单数", "商机数", "转化率（%）"]]
    
    total_task_count = 0
    total_work_order_count = 0
    total_opportunity_count = 0
    
    for stat in statistics:
        data.append([
            stat.sales_unit[:20] if len(stat.sales_unit) > 20 else stat.sales_unit,
            str(stat.task_count),
            str(stat.work_order_count),
            str(stat.opportunity_count),
            f"{stat.conversion_rate:.2f}"
        ])
        
        total_task_count += stat.task_count
        total_work_order_count += stat.work_order_count
        total_opportunity_count += stat.opportunity_count
    
    # 添加汇总行
    if len(statistics) > 0:
        total_conversion_rate = (total_opportunity_count / total_task_count * 100) if total_task_count > 0 else 0.0
        data.append([
            "合计",
            str(total_task_count),
            str(total_work_order_count),
            str(total_opportunity_count),
            f"{total_conversion_rate:.2f}"
        ])
    
    # 创建表格
    table = Table(data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C5E0B4')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        ('FONTNAME', (0, -1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # 销售单位列左对齐
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(statistics)} 个销售单位", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/statistics/sales-unit/excel")
def export_sales_unit_statistics_excel(
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出部门数据统计为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_sales_unit_statistics_to_excel(
        db, current_user, current_role, group_id, start_date_obj, end_date_obj
    )
    filename = f"部门数据统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/statistics/sales-unit/pdf")
def export_sales_unit_statistics_pdf(
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出部门数据统计为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_sales_unit_statistics_to_pdf(
        db, current_user, current_role, group_id, start_date_obj, end_date_obj
    )
    filename = f"部门数据统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


# ==================== 销售单位绩效统计导出功能 ====================

def export_sales_unit_performance_statistics_to_excel(
    db: Session, 
    current_user: User, 
    current_role,
    group_id: Optional[int] = None,
    include_member_details: bool = False,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出销售单位绩效统计为Excel（基于当前激活角色）"""
    from app.schemas.statistics import MemberDetailStatistics
    
    # 获取统计数据
    statistics, member_details = get_sales_unit_performance_statistics(
        db, current_user, current_role, group_id, include_member_details, start_date, end_date
    )
    
    # 创建Excel工作簿
    wb = Workbook()
    
    # 第一个工作表：销售单位绩效统计
    ws = wb.active
    ws.title = "销售单位绩效统计"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 数据样式
    data_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入标题
    headers = ["销售单位", "已预约", "已拜访", "有效预约率（%）", "有拜访对象权限", 
               "有效拜访率（%）", "线索数", "商机数", "线索挖掘率（%）", "线索转化率（%）"]
    for idx, header in enumerate(headers, start=1):
        col = chr(64 + idx)  # A, B, C, ...
        ws[f'{col}1'] = header
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].font = header_font
        ws[f'{col}1'].alignment = header_alignment
    
    # 写入数据
    row = 2
    for stat in statistics:
        ws[f'A{row}'] = stat.sales_unit
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
        
        ws[f'B{row}'] = stat.appointments_made
        ws[f'B{row}'].alignment = data_alignment
        
        ws[f'C{row}'] = stat.visits_completed
        ws[f'C{row}'].alignment = data_alignment
        
        ws[f'D{row}'] = stat.effective_appointment_rate
        ws[f'D{row}'].alignment = data_alignment
        ws[f'D{row}'].number_format = '0.00'
        
        ws[f'E{row}'] = stat.has_decision_authority
        ws[f'E{row}'].alignment = data_alignment
        
        ws[f'F{row}'] = stat.effective_visit_rate
        ws[f'F{row}'].alignment = data_alignment
        ws[f'F{row}'].number_format = '0.00'
        
        ws[f'G{row}'] = stat.lead_count
        ws[f'G{row}'].alignment = data_alignment
        
        ws[f'H{row}'] = stat.opportunity_count
        ws[f'H{row}'].alignment = data_alignment
        
        ws[f'I{row}'] = stat.lead_mining_rate
        ws[f'I{row}'].alignment = data_alignment
        ws[f'I{row}'].number_format = '0.00'
        
        ws[f'J{row}'] = stat.lead_conversion_rate
        ws[f'J{row}'].alignment = data_alignment
        ws[f'J{row}'].number_format = '0.00'
        
        # 总计行特殊样式
        if stat.sales_unit == "总计":
            total_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                ws[f'{col}{row}'].fill = total_fill
                ws[f'{col}{row}'].font = Font(bold=True)
        
        row += 1
    
    # 如果有成员明细，添加第二个工作表
    if member_details:
        ws2 = wb.create_sheet("成员明细统计")
        
        # 设置列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws2.column_dimensions[col].width = 15
        
        # 写入标题
        for idx, header in enumerate(headers, start=1):
            col = chr(64 + idx)
            ws2[f'{col}1'] = header.replace("销售单位", "成员")
            ws2[f'{col}1'].fill = header_fill
            ws2[f'{col}1'].font = header_font
            ws2[f'{col}1'].alignment = header_alignment
        
        # 写入成员明细数据
        row = 2
        for member in member_details:
            ws2[f'A{row}'] = member.member_name
            ws2[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
            
            ws2[f'B{row}'] = member.appointments_made
            ws2[f'B{row}'].alignment = data_alignment
            
            ws2[f'C{row}'] = member.visits_completed
            ws2[f'C{row}'].alignment = data_alignment
            
            ws2[f'D{row}'] = member.effective_appointment_rate
            ws2[f'D{row}'].alignment = data_alignment
            ws2[f'D{row}'].number_format = '0.00'
            
            ws2[f'E{row}'] = member.has_decision_authority
            ws2[f'E{row}'].alignment = data_alignment
            
            ws2[f'F{row}'] = member.effective_visit_rate
            ws2[f'F{row}'].alignment = data_alignment
            ws2[f'F{row}'].number_format = '0.00'
            
            ws2[f'G{row}'] = member.lead_count
            ws2[f'G{row}'].alignment = data_alignment
            
            ws2[f'H{row}'] = member.opportunity_count
            ws2[f'H{row}'].alignment = data_alignment
            
            ws2[f'I{row}'] = member.lead_mining_rate
            ws2[f'I{row}'].alignment = data_alignment
            ws2[f'I{row}'].number_format = '0.00'
            
            ws2[f'J{row}'] = member.lead_conversion_rate
            ws2[f'J{row}'].alignment = data_alignment
            ws2[f'J{row}'].number_format = '0.00'
            
            row += 1
    
    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_sales_unit_performance_statistics_to_pdf(
    db: Session, 
    current_user: User, 
    current_role,
    group_id: Optional[int] = None,
    include_member_details: bool = False,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出销售单位绩效统计为PDF（基于当前激活角色）"""
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    # 获取统计数据
    statistics, member_details = get_sales_unit_performance_statistics(
        db, current_user, current_role, group_id, include_member_details, start_date, end_date
    )
    
    # 创建PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=9,
    )
    
    elements.append(Paragraph("销售单位绩效统计", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 添加时间范围信息
    date_range_text = "全部时间"
    if start_date or end_date:
        date_range_text = f"{start_date.strftime('%Y-%m-%d') if start_date else ''} 至 {end_date.strftime('%Y-%m-%d') if end_date else ''}"
    elements.append(Paragraph(f"时间范围：{date_range_text}", normal_style))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 准备表格数据（简化列，因为列太多）
    data = [["销售单位", "已预约", "已拜访", "有效预约率（%）", "有决策权", 
             "有效拜访率（%）", "线索数", "商机数", "线索挖掘率（%）", "线索转化率（%）"]]
    
    for stat in statistics:
        data.append([
            stat.sales_unit[:15] if len(stat.sales_unit) > 15 else stat.sales_unit,
            str(stat.appointments_made),
            str(stat.visits_completed),
            f"{stat.effective_appointment_rate:.2f}",
            str(stat.has_decision_authority),
            f"{stat.effective_visit_rate:.2f}",
            str(stat.lead_count),
            str(stat.opportunity_count),
            f"{stat.lead_mining_rate:.2f}",
            f"{stat.lead_conversion_rate:.2f}"
        ])
    
    # 创建表格（使用较小的列宽以适应页面）
    col_widths = [1.2*inch, 0.7*inch, 0.7*inch, 0.9*inch, 0.8*inch, 0.9*inch, 0.7*inch, 0.7*inch, 0.9*inch, 0.9*inch]
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C5E0B4')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        ('FONTNAME', (0, -1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # 销售单位列左对齐
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(statistics)} 个销售单位", normal_style))
    
    # 如果有成员明细，添加成员明细表格
    if member_details:
        elements.append(Spacer(1, 0.3 * inch))
        elements.append(Paragraph("成员明细统计", ParagraphStyle(
            'SubTitle',
            parent=styles['Heading2'],
            fontName=chinese_font,
            fontSize=14,
            alignment=1,
        )))
        elements.append(Spacer(1, 0.1 * inch))
        
        member_data = [["成员", "已预约", "已拜访", "有效预约率（%）", "有决策权", 
                        "有效拜访率（%）", "线索数", "商机数", "线索挖掘率（%）", "线索转化率（%）"]]
        
        for member in member_details:
            member_data.append([
                member.member_name[:15] if len(member.member_name) > 15 else member.member_name,
                str(member.appointments_made),
                str(member.visits_completed),
                f"{member.effective_appointment_rate:.2f}",
                str(member.has_decision_authority),
                f"{member.effective_visit_rate:.2f}",
                str(member.lead_count),
                str(member.opportunity_count),
                f"{member.lead_mining_rate:.2f}",
                f"{member.lead_conversion_rate:.2f}"
            ])
        
        member_table = Table(member_data, colWidths=col_widths)
        member_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), chinese_font),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), chinese_font),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # 成员列左对齐
        ]))
        
        elements.append(member_table)
        elements.append(Spacer(1, 0.2 * inch))
        elements.append(Paragraph(f"共 {len(member_details)} 位成员", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/statistics/sales-unit-performance/excel")
def export_sales_unit_performance_statistics_excel(
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    include_member_details: bool = Query(False, description="是否包含成员明细（仅组长可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出销售单位绩效统计为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    if include_member_details and current_role.role != UserRole.TEAM_LEADER:
        raise HTTPException(status_code=403, detail="只有组长可以查看成员明细")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_sales_unit_performance_statistics_to_excel(
        db, current_user, current_role, group_id, include_member_details, start_date_obj, end_date_obj
    )
    filename = f"销售单位绩效统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/statistics/sales-unit-performance/pdf")
def export_sales_unit_performance_statistics_pdf(
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    include_member_details: bool = Query(False, description="是否包含成员明细（仅组长可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出销售单位绩效统计为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    if include_member_details and current_role.role != UserRole.TEAM_LEADER:
        raise HTTPException(status_code=403, detail="只有组长可以查看成员明细")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_sales_unit_performance_statistics_to_pdf(
        db, current_user, current_role, group_id, include_member_details, start_date_obj, end_date_obj
    )
    filename = f"销售单位绩效统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


# ==================== 线索需求方向统计导出功能 ====================

def export_requirement_direction_statistics_to_excel(
    db: Session, 
    current_user: User, 
    current_role,
    group_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出线索需求方向统计为Excel（基于当前激活角色）"""
    # 获取统计数据
    statistics = get_requirement_direction_statistics(
        db, current_user, current_role, group_id, start_date, end_date
    )
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "线索需求方向统计"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据样式
    data_alignment = Alignment(horizontal="center", vertical="center")
    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True, size=11)
    
    # 写入标题
    ws['A1'] = "分类"
    ws['B1'] = "需求方向"
    ws['C1'] = "数量"
    
    # 设置标题样式
    for col in ['A1', 'B1', 'C1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = header_alignment
    
    # 写入数据
    row = 2
    total_count = 0
    
    for group_stat in statistics:
        category = group_stat.category
        directions = group_stat.directions
        
        if not directions:
            # 单级分类（分类名就是需求方向）
            ws[f'A{row}'] = category
            ws[f'B{row}'] = category
            ws[f'C{row}'] = 0  # 如果没有方向，数量为0
            
            ws[f'A{row}'].fill = category_fill
            ws[f'A{row}'].font = category_font
            ws[f'B{row}'].fill = category_fill
            ws[f'B{row}'].font = category_font
            ws[f'C{row}'].fill = category_fill
            ws[f'C{row}'].font = category_font
            ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
            ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="center")
            ws[f'C{row}'].alignment = data_alignment
            row += 1
        else:
            # 有子方向的情况
            for idx, direction_stat in enumerate(directions):
                if idx == 0:
                    # 第一行显示分类名
                    ws[f'A{row}'] = category
                    ws[f'A{row}'].fill = category_fill
                    ws[f'A{row}'].font = category_font
                    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
                else:
                    # 后续行分类列留空（合并显示效果）
                    ws[f'A{row}'] = ""
                
                ws[f'B{row}'] = direction_stat.direction
                ws[f'B{row}'].alignment = Alignment(horizontal="left", vertical="center")
                
                ws[f'C{row}'] = direction_stat.count
                ws[f'C{row}'].alignment = data_alignment
                
                total_count += direction_stat.count
                row += 1
    
    # 添加汇总行
    if row > 2:
        total_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        ws[f'A{row}'] = "合计"
        ws[f'B{row}'] = f"共{len(statistics)}个分类"
        ws[f'C{row}'] = total_count
        
        # 设置汇总行样式
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].alignment = data_alignment
    
    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_requirement_direction_statistics_to_pdf(
    db: Session, 
    current_user: User, 
    current_role,
    group_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出线索需求方向统计为PDF（基于当前激活角色）"""
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    # 获取统计数据
    statistics = get_requirement_direction_statistics(
        db, current_user, current_role, group_id, start_date, end_date
    )
    
    # 创建PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("线索需求方向统计", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 添加时间范围信息
    date_range_text = "全部时间"
    if start_date or end_date:
        date_range_text = f"{start_date.strftime('%Y-%m-%d') if start_date else ''} 至 {end_date.strftime('%Y-%m-%d') if end_date else ''}"
    elements.append(Paragraph(f"时间范围：{date_range_text}", normal_style))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 准备表格数据
    data = [["分类", "需求方向", "数量"]]
    
    total_count = 0
    
    for group_stat in statistics:
        category = group_stat.category
        directions = group_stat.directions
        
        if not directions:
            # 单级分类
            data.append([category, category, "0"])
        else:
            # 有子方向的情况
            for idx, direction_stat in enumerate(directions):
                if idx == 0:
                    # 第一行显示分类名
                    data.append([category, direction_stat.direction, str(direction_stat.count)])
                else:
                    # 后续行分类列留空
                    data.append(["", direction_stat.direction, str(direction_stat.count)])
                
                total_count += direction_stat.count
    
    # 添加汇总行
    if len(statistics) > 0:
        data.append(["合计", f"共{len(statistics)}个分类", str(total_count)])
    
    # 创建表格
    table = Table(data, colWidths=[1.5*inch, 2.5*inch, 1*inch])
    
    # 构建样式列表
    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C5E0B4')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        ('FONTNAME', (0, -1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),  # 分类和需求方向列左对齐
    ]
    
    # 为分类行添加特殊背景色
    row_idx = 1
    for group_stat in statistics:
        directions = group_stat.directions
        if directions:
            # 第一行（分类行）使用特殊背景
            table_style.append(('BACKGROUND', (0, row_idx), (2, row_idx), colors.HexColor('#D9E1F2')))
            table_style.append(('FONTNAME', (0, row_idx), (2, row_idx), chinese_font))
            table_style.append(('FONTSIZE', (0, row_idx), (2, row_idx), 10))
            row_idx += len(directions)
        else:
            table_style.append(('BACKGROUND', (0, row_idx), (2, row_idx), colors.HexColor('#D9E1F2')))
            row_idx += 1
    
    table.setStyle(TableStyle(table_style))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(statistics)} 个分类", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/statistics/requirement-direction/excel")
def export_requirement_direction_statistics_excel(
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出线索需求方向统计为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_requirement_direction_statistics_to_excel(
        db, current_user, current_role, group_id, start_date_obj, end_date_obj
    )
    filename = f"线索需求方向统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/statistics/requirement-direction/pdf")
def export_requirement_direction_statistics_pdf(
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出线索需求方向统计为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_requirement_direction_statistics_to_pdf(
        db, current_user, current_role, group_id, start_date_obj, end_date_obj
    )
    filename = f"线索需求方向统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


# ==================== 成员工作量统计导出功能 ====================

def export_member_workloads_to_excel(
    db: Session, 
    current_user: User, 
    current_role,
    limit: int = 1000,
    group_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出成员工作量统计为Excel（基于当前激活角色）"""
    # 获取统计数据（使用较大的limit以确保导出所有数据）
    workloads = get_member_workloads(
        db, current_user, current_role, limit=limit, group_id=group_id, 
        start_date=start_date, end_date=end_date
    )
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "成员工作量统计"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # 标题样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据样式
    data_alignment = Alignment(horizontal="center", vertical="center")
    
    # 写入标题
    ws['A1'] = "成员姓名"
    ws['B1'] = "工单数量"
    ws['C1'] = "拜访日志数量"
    ws['D1'] = "已拜访工单"
    
    # 设置标题样式
    for col in ['A1', 'B1', 'C1', 'D1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = header_alignment
    
    # 写入数据
    row = 2
    total_work_orders = 0
    total_visit_logs = 0
    total_completed = 0
    
    for workload in workloads:
        ws[f'A{row}'] = workload.member_name
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
        
        ws[f'B{row}'] = workload.work_order_count
        ws[f'B{row}'].alignment = data_alignment
        
        ws[f'C{row}'] = workload.visit_log_count
        ws[f'C{row}'].alignment = data_alignment
        
        ws[f'D{row}'] = workload.completed_work_orders
        ws[f'D{row}'].alignment = data_alignment
        
        total_work_orders += workload.work_order_count
        total_visit_logs += workload.visit_log_count
        total_completed += workload.completed_work_orders
        row += 1
    
    # 添加汇总行
    if row > 2:
        total_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
        ws[f'A{row}'] = "合计"
        ws[f'B{row}'] = total_work_orders
        ws[f'C{row}'] = total_visit_logs
        ws[f'D{row}'] = total_completed
        
        # 设置汇总行样式
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = total_fill
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].alignment = data_alignment
    
    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_member_workloads_to_pdf(
    db: Session, 
    current_user: User, 
    current_role,
    limit: int = 1000,
    group_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> io.BytesIO:
    """导出成员工作量统计为PDF（基于当前激活角色）"""
    # 注册中文字体
    chinese_font = register_chinese_font()
    
    # 获取统计数据（使用较大的limit以确保导出所有数据）
    workloads = get_member_workloads(
        db, current_user, current_role, limit=limit, group_id=group_id,
        start_date=start_date, end_date=end_date
    )
    
    # 创建PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # 样式
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=chinese_font,
        fontSize=16,
        alignment=1,
    )
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=chinese_font,
        fontSize=10,
    )
    
    elements.append(Paragraph("成员工作量统计", title_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 添加时间范围信息
    date_range_text = "全部时间"
    if start_date or end_date:
        date_range_text = f"{start_date.strftime('%Y-%m-%d') if start_date else ''} 至 {end_date.strftime('%Y-%m-%d') if end_date else ''}"
    elements.append(Paragraph(f"时间范围：{date_range_text}", normal_style))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.2 * inch))
    
    # 准备表格数据
    data = [["成员姓名", "工单数量", "拜访日志数量", "已拜访工单"]]
    
    total_work_orders = 0
    total_visit_logs = 0
    total_completed = 0
    
    for workload in workloads:
        data.append([
            workload.member_name,
            str(workload.work_order_count),
            str(workload.visit_log_count),
            str(workload.completed_work_orders)
        ])
        total_work_orders += workload.work_order_count
        total_visit_logs += workload.visit_log_count
        total_completed += workload.completed_work_orders
    
    # 添加汇总行
    if len(workloads) > 0:
        data.append(["合计", str(total_work_orders), str(total_visit_logs), str(total_completed)])
    
    # 创建表格
    table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), chinese_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C5E0B4')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        ('FONTNAME', (0, -1), (-1, -1), chinese_font),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # 成员姓名列左对齐
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"共 {len(workloads)} 位成员", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/statistics/member-workload/excel")
def export_member_workloads_excel(
    limit: int = Query(1000, description="返回数量限制"),
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出成员工作量统计为Excel（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if current_role.role not in [UserRole.MANAGER, UserRole.TEAM_LEADER]:
        raise HTTPException(status_code=403, detail="只有总管和组长可以查看成员工作量统计")
    
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_member_workloads_to_excel(
        db, current_user, current_role, limit, group_id, start_date_obj, end_date_obj
    )
    filename = f"成员工作量统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )


@router.get("/statistics/member-workload/pdf")
def export_member_workloads_pdf(
    limit: int = Query(1000, description="返回数量限制"),
    group_id: Optional[int] = Query(None, description="分组ID（仅总管可用）"),
    start_date: Optional[str] = Query(None, description="开始日期（YYYY-MM-DD）"),
    end_date: Optional[str] = Query(None, description="结束日期（YYYY-MM-DD）"),
    user_role: tuple = Depends(get_current_role),
    db: Session = Depends(get_db)
):
    """导出成员工作量统计为PDF（基于当前激活角色）"""
    current_user, current_role = user_role
    
    # 权限验证
    if current_role.role not in [UserRole.MANAGER, UserRole.TEAM_LEADER]:
        raise HTTPException(status_code=403, detail="只有总管和组长可以查看成员工作量统计")
    
    if group_id and current_role.role != UserRole.MANAGER:
        raise HTTPException(status_code=403, detail="只有总管可以指定分组ID")
    
    # 解析日期参数
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="开始日期格式错误，应为YYYY-MM-DD")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="结束日期格式错误，应为YYYY-MM-DD")
    
    output = export_member_workloads_to_pdf(
        db, current_user, current_role, limit, group_id, start_date_obj, end_date_obj
    )
    filename = f"成员工作量统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filename_encoded = quote(filename, safe='')
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}; filename={filename_encoded}"
        }
    )
