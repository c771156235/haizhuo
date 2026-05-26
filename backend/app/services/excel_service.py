"""
Excel 处理服务
用于生成和解析详细需求单的 Excel 模板
"""
import re
from io import BytesIO
from typing import List, Dict, Optional
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl

DEPARTMENTS = {
    "云能力中心": None,
    "海卓": None,
    "阿网": None,
    "恒联": None,
    "政企群": None,
    "商客部": None,
    "销售单位": [
        "东区",
        "中区",
        "西区",
        "北区",
        "金山",
        "浦东",
        "公共BD",
        "互联网部",
        "政务BD",
        "科创BD（含号百）",
        "南区",
        "莘闵",
        "青浦",
        "嘉定",
        "松江",
        "数集",
        "工商BD",
        "金融BD",
        "崇明",
        "奉贤",
        "战略BD",
        "互联网BD/信网部",
        "宝山",
        "理想公司",
        "云舟",
    ],
}


def build_customer_source_options() -> List[str]:
    options: List[str] = []
    for dept, sub_depts in DEPARTMENTS.items():
        if sub_depts is None:
            options.append(dept)
        else:
            for sub_dept in sub_depts:
                options.append(f"{dept} - {sub_dept}")
    return options


def validate_contact_format(contact: str) -> bool:
    """
    验证联系方式格式（支持手机号、固定电话、邮箱）
    
    Args:
        contact: 联系方式字符串
    
    Returns:
        bool: 格式是否正确
    """
    if not contact or not contact.strip():
        return False
    
    contact = contact.strip()
    
    # 手机号格式：11位数字，1开头，第二位3-9
    phone_pattern = re.compile(r'^1[3-9]\d{9}$')
    # 固定电话格式：区号-号码（如：010-12345678, 021-12345678）
    landline_pattern = re.compile(r'^0\d{2,3}-\d{7,8}$')
    # 邮箱格式
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    return bool(phone_pattern.match(contact) or landline_pattern.match(contact) or email_pattern.match(contact))


def generate_detail_requirement_template() -> BytesIO:
    """
    生成详细需求单 Excel 模板
    
    Returns:
        BytesIO: Excel 文件的字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "详细需求单"
    
    # 设置表头
    headers = [
        "客户单位*",
        "行业类型*",
        "客户来源*",
        "详细需求内容*",
        "预期拜访时间",
        "客户拜访地址*",
        "客户经理姓名*",
        "客户经理联系方式*",
    ]
    header_row = 1
    
    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 25  # 客户单位
    ws.column_dimensions['B'].width = 20  # 行业类型
    ws.column_dimensions['C'].width = 20  # 客户来源
    ws.column_dimensions['D'].width = 50  # 详细需求内容
    ws.column_dimensions['E'].width = 20  # 预期拜访时间
    ws.column_dimensions['F'].width = 30  # 客户拜访地址
    ws.column_dimensions['G'].width = 20  # 客户经理姓名
    ws.column_dimensions['H'].width = 20  # 客户经理联系方式

    customer_source_options = build_customer_source_options()
    if customer_source_options:
        options_ws = wb.create_sheet(title="__options")
        for idx, option in enumerate(customer_source_options, start=1):
            options_ws.cell(row=idx, column=1, value=option)
        options_ws.sheet_state = "hidden"

        dv = DataValidation(
            type="list",
            formula1=f"=__options!$A$1:$A${len(customer_source_options)}",
            allow_blank=False
        )
        ws.add_data_validation(dv)
        dv.add("C2:C2000")
    
    # 添加示例数据行（可选）
    example_row = 2
    example_data = [
        "示例客户单位",
        "金融",
        "销售单位 - 东区",
        "需要AI技术支持，包括智能客服和数据分析功能",
        "2024-12-31 14:30"
    ]
    example_data.extend([
        "上海市浦东新区XX路XX号",
        "张三",
        "13800138000"
    ])
    
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    example_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=example_row, column=col_idx, value=value)
        cell.fill = example_fill
        cell.alignment = example_alignment
        cell.border = border
    
    # 设置行高
    ws.row_dimensions[header_row].height = 30
    ws.row_dimensions[example_row].height = 40
    
    # 添加说明
    ws.cell(row=example_row + 1, column=1, value="说明：")
    ws.cell(row=example_row + 1, column=1).font = Font(bold=True)
    ws.cell(row=example_row + 2, column=1, value="1. 带*号的字段为必填项")
    ws.cell(row=example_row + 3, column=1, value="2. 客户来源为必填字段，请从下拉列表选择")
    ws.cell(row=example_row + 4, column=1, value="3. 预期拜访时间格式：YYYY-MM-DD HH:mm 或 YYYY-MM-DD（如：2024-12-31 14:30 或 2024-12-31，如只填日期则时间默认为00:00）")
    ws.cell(row=example_row + 5, column=1, value="4. 客户拜访地址、客户经理姓名、客户经理联系方式为必填项")
    ws.cell(row=example_row + 6, column=1, value="5. 可以添加多行数据，每行代表一个客户的详细需求")
    ws.cell(row=example_row + 7, column=1, value="6. 上传前请删除示例行")
    # 添加重要提示：上传前必须删除说明内容
    warning_cell = ws.cell(row=example_row + 8, column=1, value="7. 【重要】上传前请删除所有说明内容（包括本行），否则系统识别会出现问题！")
    warning_cell.font = Font(bold=True, color="FF0000")  # 红色加粗字体
    
    # 将工作簿保存到字节流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def parse_detail_requirement_excel(
    file_content: bytes,
    allow_missing_customer_source: bool = False,
) -> List[Dict[str, any]]:
    """
    解析详细需求单 Excel 文件
    
    Args:
        file_content: Excel 文件的字节内容
        allow_missing_customer_source: 为 True 时允许客户来源列为空（由调用方按账号销售单位写入）

    Returns:
        List[Dict]: 解析后的详细需求列表，每个字典包含：
            - customer_unit: 客户单位
            - industry_type: 行业类型
            - customer_source: 客户来源
            - requirement_content: 详细需求内容
            - expected_visit_time: 预期拜访时间（可选）
            - customer_visit_address: 客户拜访地址（必填）
            - customer_manager_name: 客户经理姓名（必填）
            - customer_manager_contact: 客户经理联系方式（必填）
            
    Raises:
        ValueError: 如果文件格式不正确或必填字段缺失
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        ws = wb.active
        
        requirements = []
        errors = []
        
        # 从第2行开始读取数据（第1行是表头）
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not any(row):
                continue
            
            # 解析每一行
            customer_unit = str(row[0]).strip() if row[0] else ""
            industry_type = str(row[1]).strip() if row[1] else ""
            customer_source = str(row[2]).strip() if len(row) > 2 and row[2] and str(row[2]).strip() else None
            requirement_content = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            expected_visit_time_str = str(row[4]).strip() if len(row) > 4 and row[4] else None
            
            customer_visit_address = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            customer_manager_name = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            customer_manager_contact = str(row[7]).strip() if len(row) > 7 and row[7] else ""
            
            # 跳过说明行（包含"说明"、"重要"、"上传前"等关键词的行）
            if any(keyword in customer_unit for keyword in ["说明", "重要", "上传前", "删除", "示例"]):
                continue
            
            # 验证必填字段
            if not customer_unit:
                errors.append(f"第{row_idx}行：客户单位不能为空")
                continue
            if not industry_type:
                errors.append(f"第{row_idx}行：行业类型不能为空")
                continue
            if not requirement_content:
                errors.append(f"第{row_idx}行：详细需求内容不能为空")
                continue
            if not customer_source and not allow_missing_customer_source:
                errors.append(f"第{row_idx}行：客户来源不能为空")
                continue
            
            if not customer_visit_address:
                errors.append(f"第{row_idx}行：客户拜访地址不能为空")
                continue
            if not customer_manager_name:
                errors.append(f"第{row_idx}行：客户经理姓名不能为空")
                continue
            if not customer_manager_contact:
                errors.append(f"第{row_idx}行：客户经理联系方式不能为空")
                continue
            # 验证联系方式格式
            if not validate_contact_format(customer_manager_contact):
                errors.append(f"第{row_idx}行：客户经理联系方式格式不正确，请输入有效的手机号、固定电话或邮箱")
                continue
            
            # 解析日期时间
            expected_visit_time = None
            if expected_visit_time_str and expected_visit_time_str.lower() not in ['none', 'null', '']:
                try:
                    # 如果已经是日期时间对象
                    if isinstance(expected_visit_time_str, datetime):
                        expected_visit_time = expected_visit_time_str
                    elif isinstance(expected_visit_time_str, date):
                        # 如果是日期对象，转换为日期时间（默认时间为00:00:00）
                        expected_visit_time = datetime.combine(expected_visit_time_str, datetime.min.time())
                    else:
                        # 将字符串转换为字符串类型（处理可能的其他类型）
                        datetime_str = str(expected_visit_time_str).strip()
                        
                        # 尝试多种日期时间格式解析
                        parsed = False
                        datetime_formats = [
                            '%Y-%m-%d %H:%M:%S',  # 2024-12-31 14:30:00
                            '%Y/%m/%d %H:%M:%S',  # 2024/12/31 14:30:00
                            '%Y-%m-%d %H:%M',     # 2024-12-31 14:30
                            '%Y/%m/%d %H:%M',     # 2024/12/31 14:30
                            '%Y-%m-%d',           # 2024-12-31 (只有日期，时间默认为00:00:00)
                            '%Y/%m/%d',           # 2024/12/31 (只有日期，时间默认为00:00:00)
                        ]
                        
                        for fmt in datetime_formats:
                            try:
                                expected_visit_time = datetime.strptime(datetime_str, fmt)
                                parsed = True
                                break
                            except ValueError:
                                continue
                        
                        if not parsed:
                            # 解析失败，检查格式
                            import re
                            if re.match(r'^\d{4}[-/]\d{2}[-/]\d{2}', datetime_str):
                                # 格式看起来正确但日期时间无效
                                errors.append(f"第{row_idx}行：预期拜访时间无效（{expected_visit_time_str}），请检查日期时间是否正确（如：2024-12-31 14:30 或 2024-12-31）")
                            else:
                                # 格式不正确
                                errors.append(f"第{row_idx}行：预期拜访时间格式错误（{expected_visit_time_str}），应为 YYYY-MM-DD HH:mm 或 YYYY-MM-DD 格式（如：2024-12-31 14:30 或 2024-12-31）")
                            # 日期时间解析失败，但日期时间是可选的，设为 None 继续处理
                            expected_visit_time = None
                except Exception as e:
                    errors.append(f"第{row_idx}行：预期拜访时间解析失败（{expected_visit_time_str}），应为 YYYY-MM-DD HH:mm 或 YYYY-MM-DD 格式（如：2024-12-31 14:30 或 2024-12-31）")
                    # 日期时间解析失败，但日期时间是可选的，设为 None 继续处理
                    expected_visit_time = None
            
            req_dict = {
                "customer_unit": customer_unit,
                "industry_type": industry_type,
                "customer_source": customer_source,
                "requirement_content": requirement_content,
                "expected_visit_time": expected_visit_time,
                "row_number": row_idx  # 保存行号用于错误提示
            }
            
            # 添加客户来源字段（如果存在）
            if customer_source:
                req_dict["customer_source"] = customer_source
            
            req_dict["customer_visit_address"] = customer_visit_address
            req_dict["customer_manager_name"] = customer_manager_name
            req_dict["customer_manager_contact"] = customer_manager_contact
            
            requirements.append(req_dict)
        
        if errors:
            raise ValueError("Excel 文件解析错误：\n" + "\n".join(errors))
        
        if not requirements:
            raise ValueError("Excel 文件中没有有效的数据行")
        
        return requirements
        
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Excel 文件解析失败：{str(e)}")

