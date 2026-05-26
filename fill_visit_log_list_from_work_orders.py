from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = ROOT_DIR / "backend"

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

# 兼容当前环境中 DEBUG=release 之类的值，避免导入后端配置时直接失败。
# 这里强制回退到开发模式，只影响当前脚本进程，不修改项目文件。
if os.getenv("DEBUG", "").strip().lower() in {"release", "prod", "production"}:
    os.environ["DEBUG"] = "true"


DEFAULT_INPUT = ROOT_DIR / "拜访日志列表.xlsx"
DEFAULT_OUTPUT = ROOT_DIR / "拜访日志列表_已填充.xlsx"


@dataclass
class WorkOrderExportRecord:
    values_by_header: dict[str, str]


def normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def normalize_key(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def format_datetime(value: datetime | None) -> str:
    if not value:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def choose_first_non_empty(*values: Any) -> str:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def load_header_map(ws) -> dict[str, int]:
    return {
        normalize_header(cell.value): index
        for index, cell in enumerate(ws[1], start=1)
        if normalize_header(cell.value)
    }


def build_group_name_map(db) -> dict[int, str]:
    from app.models.group import Group, group_leaders

    group_name_map: dict[int, str] = {}

    for leader_id, group_name in (
        db.query(group_leaders.c.user_id, Group.name)
        .join(Group, Group.id == group_leaders.c.group_id)
        .all()
    ):
        if leader_id and group_name and leader_id not in group_name_map:
            group_name_map[int(leader_id)] = group_name

    for leader_id, group_name in db.query(Group.leader_id, Group.name).all():
        if leader_id and group_name and leader_id not in group_name_map:
            group_name_map[int(leader_id)] = group_name

    return group_name_map


def build_record_from_work_order(work_order, group_name_map: dict[int, str]) -> WorkOrderExportRecord:
    task = work_order.task
    detail_requirement: TaskDetailRequirement | None = work_order.detail_requirement
    member = work_order.member

    values_by_header = {
        "工单编号": choose_first_non_empty(work_order.work_order_no),
        "任务名称": choose_first_non_empty(getattr(task, "task_name", None)),
        "客户单位": choose_first_non_empty(
            getattr(detail_requirement, "customer_unit", None),
            getattr(task, "customer_unit", None),
        ),
        "客户拜访地址": choose_first_non_empty(
            getattr(detail_requirement, "customer_visit_address", None),
        ),
        "客户经理": choose_first_non_empty(
            getattr(detail_requirement, "customer_manager_name", None),
        ),
        "客户经理联系方式": choose_first_non_empty(
            getattr(detail_requirement, "customer_manager_contact", None),
        ),
        "组别": choose_first_non_empty(group_name_map.get(work_order.team_leader_id)),
        "所属销售单位": choose_first_non_empty(
            getattr(detail_requirement, "customer_source", None),
            getattr(task, "sales_unit", None),
        ),
        "行业": choose_first_non_empty(
            getattr(detail_requirement, "industry_type", None),
            getattr(task, "industry_type", None),
        ),
        "创建人": choose_first_non_empty(getattr(member, "real_name", None)),
        "创建时间": choose_first_non_empty(format_datetime(getattr(work_order, "created_at", None))),
    }

    return WorkOrderExportRecord(
        values_by_header=values_by_header,
    )


def apply_record_to_sheet(ws, header_map: dict[str, int], row_index: int, record: WorkOrderExportRecord) -> int:
    updated_cells = 0

    for header, value in record.values_by_header.items():
        if not value:
            continue
        column_index = header_map.get(header)
        if not column_index:
            continue
        ws.cell(row=row_index, column=column_index).value = value
        updated_cells += 1

    return updated_cells


def write_records_to_sheet(ws, header_map: dict[str, int], records: list[WorkOrderExportRecord], start_row: int = 2) -> int:
    updated_cells = 0

    for offset, record in enumerate(records):
        row_index = start_row + offset
        updated_cells += apply_record_to_sheet(ws, header_map, row_index, record)

    return updated_cells


def fetch_work_order_records(db) -> list[WorkOrderExportRecord]:
    from sqlalchemy.orm import joinedload

    from app.models.work_order import WorkOrder

    group_name_map = build_group_name_map(db)
    work_orders = (
        db.query(WorkOrder)
        .options(
            joinedload(WorkOrder.task),
            joinedload(WorkOrder.detail_requirement),
            joinedload(WorkOrder.member),
        )
        .filter(WorkOrder.status == "accepted")
        .order_by(WorkOrder.id.asc())
        .all()
    )

    return [
        build_record_from_work_order(work_order, group_name_map)
        for work_order in work_orders
        if normalize_key(work_order.work_order_no)
    ]


def fill_work_orders_to_excel(input_path: Path, output_path: Path) -> dict[str, int]:
    from app.database import SessionLocal

    workbook = load_workbook(input_path)
    worksheet = workbook[workbook.sheetnames[0]]

    header_map = load_header_map(worksheet)

    db = SessionLocal()
    try:
        records = fetch_work_order_records(db)
    finally:
        db.close()

    updated_cells = write_records_to_sheet(worksheet, header_map, records, start_row=2)

    actual_output_path = output_path
    try:
        workbook.save(actual_output_path)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        actual_output_path = output_path.with_name(
            f"{output_path.stem}_{timestamp}{output_path.suffix}"
        )
        workbook.save(actual_output_path)

    return {
        "total_records": len(records),
        "written_rows": len(records),
        "updated_cells": updated_cells,
        "output_path": str(actual_output_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据数据库中的 work_orders 数据回填拜访日志列表 Excel。"
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"输入 Excel 路径，默认：{DEFAULT_INPUT}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"输出 Excel 路径，默认：{DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = args.input.resolve()
    output_path = args.output.resolve()

    if not input_path.exists():
        print(f"未找到输入文件：{input_path}")
        return 1

    stats = fill_work_orders_to_excel(input_path, output_path)
    print(f"输入文件：{input_path}")
    print(f"输出文件：{stats['output_path']}")
    print(f"accepted 工单数：{stats['total_records']}")
    print(f"写入行数：{stats['written_rows']}")
    print(f"写入单元格数：{stats['updated_cells']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
